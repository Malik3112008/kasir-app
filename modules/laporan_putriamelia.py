from flask import render_template, request, redirect, url_for, session, Response
from database import db
from helpers import format_kbbi_date, formatRp
from modules.blueprints import admin_bp
import io

@admin_bp.route('/admin/cetak_laporan', methods=['GET', 'POST'])
def admin_cetak_laporan():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))

    tanggal_awal = request.args.get('tanggal_awal', '')
    tanggal_akhir = request.args.get('tanggal_akhir', '')

    data_transaksi = []
    for p in db.pesanan:
        if (not tanggal_awal or p['tanggal'] >= tanggal_awal) and (not tanggal_akhir or p['tanggal'] <= tanggal_akhir):
            total_barang = sum(b['harga'] * b['jumlah'] for b in p['barang'])
            data_transaksi.append({
                'tanggal': p['tanggal'],
                'id': p['id'],
                'jumlah_item': sum(b['jumlah'] for b in p['barang']),
                'total': total_barang
            })

    if not data_transaksi and not (tanggal_awal or tanggal_akhir):
        data_transaksi = [
            {'tanggal': '2026-05-01', 'id': 'TRX001', 'jumlah_item': 3, 'total': 15000},
            {'tanggal': '2026-05-02', 'id': 'TRX002', 'jumlah_item': 5, 'total': 25000},
            {'tanggal': '2026-05-03', 'id': 'TRX003', 'jumlah_item': 2, 'total': 10000},
        ]

    total_pendapatan = sum(t['total'] for t in data_transaksi)
    modal_barang = int(total_pendapatan * 0.7)
    untung_rugi = total_pendapatan - modal_barang

    return render_template('12.-cetaklaporan.html',
        barang=db.data_barang, formatRp=formatRp,
        total_pendapatan=total_pendapatan,
        modal_barang=modal_barang,
        untung_rugi=untung_rugi,
        data_transaksi=data_transaksi,
        data_barang_list=db.data_barang,
        pesanan_list=db.pesanan,
        tanggal_awal=tanggal_awal,
        tanggal_akhir=tanggal_akhir)

@admin_bp.route('/admin/cetak_transaksi_pdf')
def admin_cetak_transaksi_pdf():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    tanggal_awal = request.args.get('tanggal_awal', '')
    tanggal_akhir = request.args.get('tanggal_akhir', '')
    filtered_pesanan = []
    for p in db.pesanan:
        if (not tanggal_awal or p['tanggal'] >= tanggal_awal) and (not tanggal_akhir or p['tanggal'] <= tanggal_akhir):
            filtered_pesanan.append(p)
    return render_template('12.-cetaklaporan_transaksi_pdf.html', pesanan=filtered_pesanan)

@admin_bp.route('/admin/cetak_transaksi_excel')
def admin_cetak_transaksi_excel():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    tanggal_awal = request.args.get('tanggal_awal', '')
    tanggal_akhir = request.args.get('tanggal_akhir', '')
    filtered_pesanan = []
    for p in db.pesanan:
        if (not tanggal_awal or p['tanggal'] >= tanggal_awal) and (not tanggal_akhir or p['tanggal'] <= tanggal_akhir):
            filtered_pesanan.append(p)
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Transaksi"
    ws.append(['No', 'Tanggal', 'ID Transaksi', 'Nama Produk', 'Total', 'Metode', 'Status'])
    for i, p in enumerate(filtered_pesanan, 1):
        total_barang = sum(b['harga'] * b['jumlah'] for b in p['barang'])
        nama_produk = ', '.join(b['nama'] for b in p['barang'])
        tgl_fmt = format_kbbi_date(p['tanggal'])
        ws.append([i, tgl_fmt, p['id'], nama_produk, formatRp(total_barang), p['metode'], p['status']])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment;filename=laporan_transaksi.xlsx'}
    )

@admin_bp.route('/admin/pengaturan-laporan')
def admin_pengaturan_laporan():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    tanggal_awal = request.args.get('tanggal_awal', '')
    tanggal_akhir = request.args.get('tanggal_akhir', '')
    laporan_type = request.args.get('laporan', 'barang')
    return render_template("16.pengaturan_laporan.html",
                           barang=db.data_barang,
                           tanggal_awal=tanggal_awal,
                           tanggal_akhir=tanggal_akhir,
                           laporan_type=laporan_type)

@admin_bp.route('/admin/cetak_laporan_barang')
def admin_cetak_laporan_barang():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    total_nilai = sum(b['harga'] * b['stok'] for b in db.data_barang)
    return render_template('12.-cetaklaporan_barang.html',
        data_barang=db.data_barang,
        total_nilai=total_nilai)

@admin_bp.route('/admin/cetak_barang_pdf')
def admin_cetak_barang_pdf():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    tanggal_awal = request.args.get('tanggal_awal', '')
    tanggal_akhir = request.args.get('tanggal_akhir', '')
    filtered_barang = []
    for b in db.data_barang:
        if (not tanggal_awal or b['tanggal'] >= tanggal_awal) and (not tanggal_akhir or b['tanggal'] <= tanggal_akhir):
            filtered_barang.append(b)
    total_nilai = sum(b['harga'] * b['stok'] for b in filtered_barang)
    return render_template('12.-cetaklaporan_barang_pdf.html',
        data_barang=filtered_barang,
        total_nilai=total_nilai)

@admin_bp.route('/admin/cetak_barang_excel')
def admin_cetak_barang_excel():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    tanggal_awal = request.args.get('tanggal_awal', '')
    tanggal_akhir = request.args.get('tanggal_akhir', '')
    filtered_barang = []
    for b in db.data_barang:
        if (not tanggal_awal or b['tanggal'] >= tanggal_awal) and (not tanggal_akhir or b['tanggal'] <= tanggal_akhir):
            filtered_barang.append(b)
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Data Barang"
    ws.append(['No', 'Nama', 'Stok', 'Harga', 'Kategori', 'Tanggal', 'Total'])
    for b in filtered_barang:
        ws.append([b['no'], b['nama'], b['stok'], formatRp(b['harga']), b['kategori'], format_kbbi_date(b['tanggal']), formatRp(b['harga'] * b['stok'])])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment;filename=laporan_data_barang.xlsx'}
    )

@admin_bp.route('/admin/cetak_pdf')
def admin_cetak_pdf():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))

    tanggal_awal = request.args.get('tanggal_awal', '')
    tanggal_akhir = request.args.get('tanggal_akhir', '')

    data_transaksi = []
    for p in db.pesanan:
        if (not tanggal_awal or p['tanggal'] >= tanggal_awal) and (not tanggal_akhir or p['tanggal'] <= tanggal_akhir):
            total_barang = sum(b['harga'] * b['jumlah'] for b in p['barang'])
            data_transaksi.append({
                'tanggal': p['tanggal'],
                'id': p['id'],
                'jumlah': sum(b['jumlah'] for b in p['barang']),
                'total': total_barang
            })

    if not data_transaksi and not (tanggal_awal or tanggal_akhir):
        data_transaksi = [
            {'tanggal': '2026-05-01', 'id': 'TRX001', 'jumlah': 3, 'total': 15000},
            {'tanggal': '2026-05-02', 'id': 'TRX002', 'jumlah': 5, 'total': 25000},
            {'tanggal': '2026-05-03', 'id': 'TRX003', 'jumlah': 2, 'total': 10000},
        ]

    total_pendapatan = sum(t['total'] for t in data_transaksi)
    modal_barang = int(total_pendapatan * 0.7)
    untung_rugi = total_pendapatan - modal_barang

    return render_template('12.-cetaklaporan_pdf.html',
        total_pendapatan=total_pendapatan,
        modal_barang=modal_barang,
        untung_rugi=untung_rugi,
        data_transaksi=data_transaksi)

@admin_bp.route('/admin/cetak_excel')
def admin_cetak_excel():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))

    tanggal_awal = request.args.get('tanggal_awal', '')
    tanggal_akhir = request.args.get('tanggal_akhir', '')

    data_transaksi = []
    for p in db.pesanan:
        if (not tanggal_awal or p['tanggal'] >= tanggal_awal) and (not tanggal_akhir or p['tanggal'] <= tanggal_akhir):
            total_barang = sum(b['harga'] * b['jumlah'] for b in p['barang'])
            data_transaksi.append({
                'tanggal': p['tanggal'],
                'id': p['id'],
                'jumlah': sum(b['jumlah'] for b in p['barang']),
                'total': total_barang
            })

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Penjualan"
    ws.append(['No', 'Tanggal', 'ID Transaksi', 'Jumlah Item', 'Total Pendapatan'])
    for i, t in enumerate(data_transaksi, 1):
        ws.append([i, format_kbbi_date(t['tanggal']), t['id'], t['jumlah'], formatRp(t['total'])])
    
    total_pendapatan = sum(t['total'] for t in data_transaksi)
    modal_barang = int(total_pendapatan * 0.7)
    untung_rugi = total_pendapatan - modal_barang

    ws.append([])
    ws.append(['Total Pendapatan', formatRp(total_pendapatan)])
    ws.append(['Modal Barang (70%)', formatRp(modal_barang)])
    ws.append(['Untung/Rugi', formatRp(untung_rugi)])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment;filename=laporan_penjualan.xlsx'}
    )
