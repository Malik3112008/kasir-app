from flask import Flask, render_template, request, redirect, url_for, session, make_response, Response, jsonify
from werkzeug.utils import secure_filename
from jinja2 import ChoiceLoader, FileSystemLoader
import os
import secrets
import json
import csv
import io
import random
from datetime import datetime

# ============================================================
# APP SETUP
# ============================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__, static_folder=os.path.join(BASE_DIR, 'static'), static_url_path='/static')
app.secret_key = os.environ.get('FLASK_SECRET', 'dev_secret_key')
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0  # No caching during dev, change to 300 for prod

# Load templates dari kasir-admin dan kasir-pembeli
app.jinja_loader = ChoiceLoader([
    FileSystemLoader(os.path.join(BASE_DIR, 'kasir-admin', 'templates')),
    FileSystemLoader(os.path.join(BASE_DIR, 'kasir-pembeli', 'templates')),
])


def format_kbbi_date(val):
    if not val:
        return ""
    from datetime import datetime, date
    tgl = None
    if isinstance(val, (datetime, date)):
        tgl = val
    else:
        val_str = str(val).strip()
        if not val_str or val_str == '-':
            return val_str
        
        # Try various formats
        formats = [
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%d %H:%M',
            '%d-%m-%Y %H:%M',
            '%d-%m-%Y %H:%M:%S',
            '%d/%m/%Y %H:%M',
            '%d/%m/%Y %H:%M:%S',
            '%Y-%m-%d',
            '%d-%m-%Y',
            '%d/%m/%Y'
        ]
        for fmt in formats:
            try:
                tgl = datetime.strptime(val_str, fmt)
                break
            except ValueError:
                continue
                
        if not tgl:
            # Clean textual dates like "28 Nov 2025" or "10 Agustus 2025"
            month_map = {
                'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'mei': 5, 'may': 5, 'jun': 6, 'jul': 7,
                'agu': 8, 'aug': 8, 'agt': 8, 'sep': 9, 'okt': 10, 'oct': 10, 'nov': 11, 'des': 12, 'dec': 12,
                'januari': 1, 'februari': 2, 'maret': 3, 'april': 4, 'juni': 6, 'juli': 7, 'agustus': 8,
                'september': 9, 'oktober': 10, 'november': 11, 'desember': 12, 'january': 1, 'february': 2,
                'march': 3, 'june': 6, 'july': 7, 'august': 8, 'october': 10, 'december': 12
            }
            import re
            m = re.search(r'(\d{1,2})\s+([a-zA-Z]+)\s+(\d{4})(?:\s+(\d{1,2})[:\.](\d{1,2}))?', val_str)
            if m:
                day = int(m.group(1))
                mon_str = m.group(2).lower()
                year = int(m.group(3))
                hour = m.group(4)
                minute = m.group(5)
                if mon_str in month_map:
                    mon = month_map[mon_str]
                    if hour and minute:
                        tgl = datetime(year, mon, day, int(hour), int(minute))
                    else:
                        tgl = datetime(year, mon, day)
            
            if not tgl:
                return val_str

    has_time = False
    if isinstance(val, datetime):
        if val.hour != 0 or val.minute != 0 or val.second != 0:
            has_time = True
    elif isinstance(val, str):
        val_str = str(val).strip()
        if ' ' in val_str and (':' in val_str or '.' in val_str):
            has_time = True
            
    if has_time and hasattr(tgl, 'hour'):
        return tgl.strftime('%d-%m-%Y %H:%M')
    else:
        return tgl.strftime('%d-%m-%Y')

@app.template_filter('tgl_indo')
def tgl_indo_filter(tanggal_str):
    return format_kbbi_date(tanggal_str)

@app.context_processor
def inject_now():
    from datetime import datetime
    return {
        'now': format_kbbi_date(datetime.now())
    }


# ============================================================
# PERSISTENCE HELPER FUNCTIONS & API ENDPOINTS
# ============================================================

BARANG_FILE = os.path.join(BASE_DIR, 'data_barang.json')
PESANAN_FILE = os.path.join(BASE_DIR, 'pesanan.json')

def hitung_total_barang(barang):
    total = 0
    for b in barang:
        total += b["jumlah"] * b["harga"]
    return total

def load_data_barang():
    if os.path.exists(BARANG_FILE):
        try:
            with open(BARANG_FILE, 'r') as f:
                return json.load(f)
        except Exception as e:
            print(f"Error loading {BARANG_FILE}: {e}")
    # static fallback
    return [
        {"no": 1, "nama": "Roti Aoka", "berat": "60", "satuan": "gr", "stok": 15, "harga": 3000, "kategori": "Makanan", "tanggal_restok": "2025-09-25", "expired": "Mei 2028", "tanggal": "2026-05-06", "gambar": "gambar-dan-icon/gambar-roti-aoka.jpeg", "rating": 5, "emoji": "🍞"},
        {"no": 2, "nama": "Donat", "berat": "50", "satuan": "gr", "stok": 10, "harga": 5000, "kategori": "Makanan", "tanggal_restok": "2025-10-01", "expired": "Mei 2028", "tanggal": "2026-05-05", "gambar": "gambar-dan-icon/donat.jpg", "rating": 4, "emoji": "🍩"},
        {"no": 3, "nama": "Pop Mie", "berat": "85", "satuan": "gr", "stok": 50, "harga": 4000, "kategori": "Makanan", "tanggal_restok": "2025-08-10", "expired": "Mei 2028", "tanggal": "2026-05-03", "gambar": "gambar-dan-icon/pop-mie.png", "rating": 4, "emoji": "🍜"},
        {"no": 4, "nama": "Air Mineral", "berat": "600", "satuan": "ml", "stok": 35, "harga": 3000, "kategori": "Minuman", "tanggal_restok": "2025-11-03", "expired": "Mei 2028", "tanggal": "2026-05-03", "gambar": "gambar-dan-icon/ades.jpg", "rating": 5, "emoji": "💧"},
        {"no": 5, "nama": "Teh Botol", "berat": "350", "satuan": "ml", "stok": 18, "harga": 5000, "kategori": "Minuman", "tanggal_restok": "2025-06-01", "expired": "Desember 2027", "tanggal": "2026-05-10", "gambar": "gambar-dan-icon/teh-botol.png", "rating": 3, "emoji": "🥤"},
        {"no": 6, "nama": "Ultra Milk", "berat": "250", "satuan": "ml", "stok": 12, "harga": 7000, "kategori": "Minuman", "tanggal_restok": "2025-06-01", "expired": "Desember 2027", "tanggal": "2026-05-10", "gambar": "gambar-dan-icon/ultramilk.png", "rating": 4, "emoji": "🧃"},
        {"no": 7, "nama": "Pensil", "berat": "10", "satuan": "gr", "stok": 0, "harga": 2000, "kategori": "Alat Tulis", "tanggal_restok": "2025-06-01", "expired": "-", "tanggal": "2026-05-10", "gambar": "gambar-dan-icon/gambar-pensil.jpeg", "rating": 4, "emoji": "✏️"},
        {"no": 8, "nama": "Buku Tulis", "berat": "100", "satuan": "gr", "stok": 0, "harga": 6000, "kategori": "Alat Tulis", "tanggal_restok": "2025-06-01", "expired": "-", "tanggal": "2026-05-10", "gambar": "gambar-dan-icon/buku-tulis.jpg", "rating": 4, "emoji": "📓"},
        {"no": 9, "nama": "Penghapus", "berat": "20", "satuan": "gr", "stok": 15, "harga": 2000, "kategori": "Alat Tulis", "tanggal_restok": "2025-06-01", "expired": "-", "tanggal": "2026-05-10", "gambar": "gambar-dan-icon/penghapus.png", "rating": 5, "emoji": "🧼"},
        {"no": 10, "nama": "Bolpoin", "berat": "30", "satuan": "gr", "stok": 14, "harga": 8000, "kategori": "Alat Tulis", "tanggal_restok": "2025-06-01", "expired": "-", "tanggal": "2026-05-10", "gambar": "gambar-dan-icon/bulpoin.jpg", "rating": 4, "emoji": "🖋️"}
    ]

def save_data_barang(data):
    try:
        with open(BARANG_FILE, 'w') as f:
            json.dump(data, f, indent=4)
    except Exception as e:
        print(f"Error saving {BARANG_FILE}: {e}")

def load_pesanan():
    data = []
    if os.path.exists(PESANAN_FILE):
        try:
            with open(PESANAN_FILE, 'r') as f:
                data = json.load(f)
        except Exception as e:
            print(f"Error loading {PESANAN_FILE}: {e}")
            data = get_default_pesanan()
    else:
        data = get_default_pesanan()
        
    for p in data:
        if "total_awal" not in p:
            p["total_awal"] = hitung_total_barang(p["barang"])
        if "total" not in p:
            p["total"] = p["total_awal"]
        if "refund" not in p:
            p["refund"] = 0
    return data

def get_default_pesanan():
    return [
        {"id": "TRX001", "tanggal": "2025-11-15", "pelanggan": "Ahmad Rizki", "metode": "Tunai", "status": "Disiapkan",
         "barang": [{"nama": "Sabun", "jumlah": 2, "harga": 5000, "gambar": "gambar-dan-icon/sabun.jpg"}, {"nama": "Pop Mie", "jumlah": 3, "harga": 4000, "gambar": "gambar-dan-icon/pop-mie.png"}, {"nama": "Air Mineral", "jumlah": 1, "harga": 3000, "gambar": "gambar-dan-icon/ades.jpg"}]},
        {"id": "TRX002", "tanggal": "2025-11-15", "pelanggan": "Siti Nurhazila", "metode": "QRIS", "status": "Disiapkan",
         "barang": [{"nama": "Penggaris", "jumlah": 1, "harga": 4000, "gambar": "gambar-dan-icon/gambar-pensil.jpeg"}, {"nama": "Penghapus", "jumlah": 2, "harga": 3000, "gambar": "gambar-dan-icon/gambar-penghapus.jpeg"}, {"nama": "Pensil", "jumlah": 4, "harga": 10000, "gambar": "gambar-dan-icon/gambar-pensil.jpeg"}]},
        {"id": "TRX003", "tanggal": "2025-11-15", "pelanggan": "Diki Nurhazila", "metode": "Tunai", "status": "Disiapkan",
         "barang": [{"nama": "Penggaris", "jumlah": 1, "harga": 4000, "gambar": "gambar-dan-icon/gambar-pensil.jpeg"}, {"nama": "Penghapus", "jumlah": 2, "harga": 3000, "gambar": "gambar-dan-icon/gambar-penghapus.jpeg"}, {"nama": "Pensil", "jumlah": 4, "harga": 10000, "gambar": "gambar-dan-icon/gambar-pensil.jpeg"}]},
        {"id": "TRX004", "tanggal": "2025-11-15", "pelanggan": "Siti riki", "metode": "QRIS", "status": "Disiapkan",
         "barang": [{"nama": "Penggaris", "jumlah": 1, "harga": 4000, "gambar": "gambar-dan-icon/gambar-pensil.jpeg"}, {"nama": "Penghapus", "jumlah": 2, "harga": 3000, "gambar": "gambar-dan-icon/gambar-penghapus.jpeg"}, {"nama": "Pensil", "jumlah": 4, "harga": 10000, "gambar": "gambar-dan-icon/gambar-pensil.jpeg"}]}
    ]

def save_pesanan(data):
    try:
        with open(PESANAN_FILE, 'w') as f:
            json.dump(data, f, indent=4)
    except Exception as e:
        print(f"Error saving {PESANAN_FILE}: {e}")

@app.before_request
def load_db_to_globals():
    global data_barang, pesanan
    data_barang = load_data_barang()
    pesanan = load_pesanan()

# Expose REST API endpoints
@app.route('/api/barang', methods=['GET', 'POST'])
def api_barang():
    global data_barang
    if request.method == 'GET':
        return jsonify(data_barang)
    elif request.method == 'POST':
        req_data = request.get_json() or request.form
        if not req_data:
            return jsonify({"error": "No data provided"}), 400
        no_baru = max([b['no'] for b in data_barang], default=0) + 1
        new_item = {
            'no': no_baru,
            'nama': req_data.get('nama', ''),
            'berat': req_data.get('berat', '-'),
            'satuan': req_data.get('satuan', ''),
            'stok': int(req_data.get('stok', 0)),
            'harga': int(req_data.get('harga', 0)),
            'kategori': req_data.get('kategori', ''),
            'tanggal_restok': req_data.get('tanggal_restok', ''),
            'expired': req_data.get('expired', '-'),
            'tanggal': req_data.get('tanggal', datetime.now().strftime("%Y-%m-%d")),
            'gambar': req_data.get('gambar', ''),
            'rating': int(req_data.get('rating', 0)),
            'emoji': req_data.get('emoji', '📦')
        }
        data_barang.append(new_item)
        save_data_barang(data_barang)
        return jsonify(new_item), 201

@app.route('/api/barang/<int:id>', methods=['GET', 'PUT', 'DELETE'])
def api_barang_detail(id):
    global data_barang
    item = next((b for b in data_barang if b['no'] == id), None)
    if not item:
        return jsonify({"error": "Product not found"}), 404
        
    if request.method == 'GET':
        return jsonify(item)
    elif request.method == 'PUT':
        req_data = request.get_json() or request.form
        if not req_data:
            return jsonify({"error": "No data provided"}), 400
        item['nama'] = req_data.get('nama', item['nama'])
        item['berat'] = req_data.get('berat', item['berat'])
        item['satuan'] = req_data.get('satuan', item.get('satuan', ''))
        item['stok'] = int(req_data.get('stok', item['stok']))
        item['harga'] = int(req_data.get('harga', item['harga']))
        item['kategori'] = req_data.get('kategori', item['kategori'])
        item['tanggal_restok'] = req_data.get('tanggal_restok', item.get('tanggal_restok', ''))
        item['expired'] = req_data.get('expired', item.get('expired', ''))
        item['gambar'] = req_data.get('gambar', item.get('gambar', ''))
        item['rating'] = int(req_data.get('rating', item.get('rating', 0)))
        item['emoji'] = req_data.get('emoji', item.get('emoji', '📦'))
        save_data_barang(data_barang)
        return jsonify(item)
    elif request.method == 'DELETE':
        data_barang = [b for b in data_barang if b['no'] != id]
        save_data_barang(data_barang)
        return jsonify({"message": "Product deleted successfully"})

@app.route('/api/pesanan', methods=['GET'])
def api_pesanan():
    global pesanan
    return jsonify(pesanan)

# ============================================================

# ============================================================
# 404 NOT FOUND
# ============================================================

@app.errorhandler(404)
def admin_beranda_awal(e):
    return render_template('03.Beranda_awal.html')

# ============================================================
# ADMIN: BERANDA AWAL
# ============================================================

@app.route('/admin')
def admin_beranda_awal():
    return render_template('03.Beranda_awal.html')

@app.route('/informasi')
def informasi():
    try:
        with open(os.path.join(BASE_DIR, 'data_koperasi.json'), 'r') as f:
            koperasi = json.load(f)
    except:
        koperasi = {}
    return render_template('informasi.html', koperasi=koperasi)

# ============================================================
# ADMIN: DENAH
# ============================================================

CARDS_DATA = [
    {"id": 1, "text": "Rak Makanan Ringan", "icon": "fa-solid fa-cookie", "href": "/admin/denah/makanan_ringan", "width": 300, "height": 160, "left": 100, "top": 80, "image": "images/makanan_ringan/gambar1.jpg"},
    {"id": 2, "text": "Rak Snack", "icon": "fa-solid fa-candy-cane", "href": "/admin/denah/snack", "width": 300, "height": 160, "left": 100, "top": 300, "image": "images/snack/gambar1.jpg"},
    {"id": 3, "text": "Rak Cemilan", "icon": "fa-solid fa-cookie-bite", "href": "/admin/denah/cemilan", "width": 300, "height": 160, "left": 100, "top": 520, "image": "images/cemilan/gambar1.jpg"},
    {"id": 4, "text": "Meja Kasir", "icon": "fa-solid fa-cash-register", "href": "/admin/denah/meja_kasir", "width": 250, "height": 220, "left": 550, "top": 80, "image": "images/meja_kasir/gambar1.jpg"},
    {"id": 5, "text": "Rak Alat Tulis", "icon": "fa-solid fa-book", "href": "/admin/denah/alat_tulis", "width": 200, "height": 300, "left": 950, "top": 80, "image": "images/alat_tulis/gambar1.jpg"},
    {"id": 6, "text": "Rak Makanan", "icon": "fa-solid fa-bowl-food", "href": "/admin/denah/makanan", "width": 200, "height": 300, "left": 950, "top": 420, "image": "images/makanan/gambar1.jpg"},
    {"id": 7, "text": "Rak Minuman", "icon": "fa-solid fa-bottle-water", "href": "/admin/denah/minuman", "width": 180, "height": 650, "left": 1250, "top": 80, "image": "images/minuman/gambar1.jpg"},
]

def load_cards():
    return CARDS_DATA

def save_cards(cards):
    global CARDS_DATA
    CARDS_DATA = cards

@app.route('/admin/denah')
def admin_denah():
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    if session.get('role') != 'admin':
        return redirect('/pembeli/denah')
    cards = load_cards()
    is_admin = True
    edit_mode = request.args.get('edit') == '1'
    edit_card_id = request.args.get('edit_card', type=int) if edit_mode else None
    edit_card = None
    if edit_card_id is not None:
        for card in cards:
            if card['id'] == edit_card_id:
                edit_card = card
                break
    return render_template('04.Denah.html', cards=cards, edit_mode=edit_mode, edit_card=edit_card, is_admin=is_admin, page='home')

@app.route('/admin/delete/<int:card_id>', methods=['POST'])
def admin_delete_card(card_id):
    if not session.get('user'):
        return "Unauthorized", 403
    cards = load_cards()
    cards = [c for c in cards if c['id'] != card_id]
    save_cards(cards)
    return redirect('/admin/denah?edit=1')

@app.route('/admin/update/<int:card_id>', methods=['POST'])
def admin_update_card(card_id):
    if not session.get('user'):
        return "Unauthorized", 403
    cards = load_cards()
    for card in cards:
        if card['id'] == card_id:
            card['text'] = request.form.get('text', card['text'])
            card['icon'] = request.form.get('icon', card['icon'])
            card['icon_size'] = int(request.form.get('icon_size', card.get('icon_size', 30)))
            card['width'] = int(request.form.get('width', card['width']))
            card['height'] = int(request.form.get('height', card['height']))
            break
    save_cards(cards)
    return redirect('/admin/denah?edit=1')

@app.route('/admin/move/<int:card_id>', methods=['POST'])
def admin_move_card(card_id):
    if not session.get('user'):
        return "Unauthorized", 403
    data = request.get_json()
    cards = load_cards()
    for card in cards:
        if card['id'] == card_id:
            card['left'] = int(data.get('left', card['left']))
            card['top'] = int(data.get('top', card['top']))
            break
    save_cards(cards)
    return jsonify({'ok': True})

@app.route('/dynamic_cards.css')
def dynamic_cards_css():
    cards = load_cards()
    css_content = ""
    for card in cards:
        css_content += f".card-id-{card['id']} {{ width: {card['width']}px; height: {card['height']}px; left: {card['left']}px; top: {card['top']}px; }}\n"
    response = make_response(css_content)
    response.headers['Content-Type'] = 'text/css'
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return response

@app.route('/admin/denah/<folder>')
def admin_detail(folder):
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    if session.get('role') != 'admin':
        return redirect(url_for('pembeli_detail', folder=folder))
    title = folder.replace('-', ' ').replace('_', ' ').title()
    return render_template('04.Denah.html', title=title, folder=folder, page='detail', is_admin=True)

@app.route('/pembeli/denah')
def pembeli_denah():
    if not session.get('user'):
        return redirect(url_for('pembeli_login'))
    cards = load_cards()
    is_admin = False
    edit_mode = False
    return render_template('04.Denah.html', cards=cards, edit_mode=edit_mode, edit_card=None, is_admin=is_admin, page='home')

@app.route('/pembeli/denah/<folder>')
def pembeli_detail(folder):
    if not session.get('user'):
        return redirect(url_for('pembeli_login'))
    title = folder.replace('-', ' ').replace('_', ' ').title()
    return render_template('04.Denah.html', title=title, folder=folder, page='detail', is_admin=False)

# ============================================================
# ADMIN: LOGIN, REGISTER, FORGOT PASSWORD
# ============================================================

USERS = {'admin': 'admin123', 'pembeli': 'beli123'}
EMAIL_TO_USER = {}

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        if USERS.get(username) == password:
            session['user'] = username
            session['role'] = 'admin'
            return redirect(url_for('admin_dashboard'))
        error = 'Nama akun atau kata sandi tidak cocok.'
    return render_template('05.1.login.html', error=error)

@app.route('/admin/logout')
def admin_logout():
    session.clear()
    return redirect(url_for('admin_login'))

@app.route('/admin/register', methods=['GET', 'POST'])
def admin_register():
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        confirm = request.form.get('confirm', '')
        if not username or not email or not password:
            error = 'Lengkapi semua bidang.'
        elif password != confirm:
            error = 'Kata sandi dan konfirmasi tidak cocok.'
        elif username in USERS:
            error = 'Nama pengguna sudah terdaftar.'
        else:
            USERS[username] = password
            EMAIL_TO_USER[email] = username
            return redirect(url_for('admin_login'))
    return render_template('05.2.register.html', error=error)

@app.route('/admin/forgot', methods=['GET', 'POST'])
def admin_forgot():
    error = None
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        session['reset_email'] = email
        session['reset_user'] = EMAIL_TO_USER.get(email)
        otp = f"{secrets.randbelow(900000) + 100000}"
        session['otp'] = otp
        print(f"[DEBUG] OTP for {email}: {otp}")
        return redirect(url_for('admin_verify'))
    return render_template('05.3.forgot.html', error=error)

@app.route('/admin/verify', methods=['GET', 'POST'])
def admin_verify():
    error = None
    if request.method == 'POST':
        otp = request.form.get('otp', '').strip()
        if otp and otp == session.get('otp'):
            return redirect(url_for('admin_reset'))
        error = 'Kode OTP tidak valid.'
    return render_template('05.4.verify_otp.html', error=error, email=session.get('reset_email'))

@app.route('/admin/reset', methods=['GET', 'POST'])
def admin_reset():
    error = None
    if request.method == 'POST':
        password = request.form.get('password', '')
        confirm = request.form.get('confirm', '')
        if not password:
            error = 'Masukkan kata sandi baru.'
        elif password != confirm:
            error = 'Konfirmasi kata sandi tidak cocok.'
        else:
            user = session.get('reset_user')
            if user:
                USERS[user] = password
            session.pop('otp', None)
            session.pop('reset_email', None)
            session.pop('reset_user', None)
            return redirect(url_for('admin_login'))
    return render_template('05.5.reset.html', error=error)

# ============================================================
# ADMIN: NOTIFIKASI
# ============================================================

notifikasi = [
    {"judul": "Pembelian Baru", "isi": "Transaksi #TRX005 oleh Ahmad Rizki senilai 18.000", "waktu": "10 menit lalu", "warna": "biru"},
    {"judul": "Stok Menipis", "isi": "Keripik singkong tersisa 3 unit", "waktu": "15 menit lalu", "warna": "orange"},
    {"judul": "Pembayaran Dikonfirmasi", "isi": "Pembayaran QRIS oleh Putri Amel senilai 10.000 telah divalidasi", "waktu": "20 menit lalu", "warna": "hijau"},
]

riwayat = []

@app.route("/admin/notifikasi")
def admin_notifikasi():
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    return render_template("06.NotifikasiAdmin.html", notifikasi=notifikasi)

@app.route("/admin/riwayat")
def admin_riwayat():
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    return render_template("06.RiwayatTransaksi.html", pesanan=pesanan)

@app.route("/admin/detail-transaksi/<trx_id>")
def admin_detail_transaksi(trx_id):
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    trx = None
    for p in pesanan:
        if p['id'] == trx_id:
            trx = p
            break
    if not trx:
        return "Transaksi tidak ditemukan", 404
    items = []
    for b in trx['barang']:
        items.append({
            'nama': b['nama'],
            'harga': b['harga'],
            'jumlah': b['jumlah'],
            'id_transaksi': trx['id'],
            'gambar': b.get('gambar', '')
        })
    total = sum(b['harga'] * b['jumlah'] for b in trx['barang'])
    tanggal_fmt = format_kbbi_date(trx['tanggal'])
    return render_template("detail_transaksi.html", items=items, total=total, tanggal=tanggal_fmt)

# ============================================================
# ADMIN: RIWAYAT AKTIVITAS
# ============================================================

data_aktivitas = [
    {"tipe": "login", "catatan": "Log in system", "status": "Sukses", "admin": "Admin", "waktu": "2026-05-28 08:00"},
    {"tipe": "tambah", "catatan": "Menambahkan pilihan barang: Roti Aoka", "status": "Berhasil", "admin": "Admin", "waktu": "2026-05-28 07:30"},
    {"tipe": "restok", "catatan": "Melakukan restok produk: Air Mineral (+50 unit)", "status": "Berhasil", "admin": "Admin", "waktu": "2026-05-28 07:00"},
    {"tipe": "logout", "catatan": "Log out system", "status": "Sukses", "admin": "Admin", "waktu": "2026-05-27 17:00"},
    {"tipe": "ubah", "catatan": "Mengubah harga: Bolpoin (Rp 3.500 -> Rp 4.000)", "status": "Berhasil", "admin": "Admin", "waktu": "2026-05-27 14:30"},
    {"tipe": "hapus", "catatan": "Menghapus pilihan barang: Kerupuk Bawang", "status": "Berhasil", "admin": "Admin", "waktu": "2026-05-27 11:20"},
    {"tipe": "login", "catatan": "Log in system", "status": "Sukses", "admin": "Admin", "waktu": "2026-05-27 08:00"},
]

@app.route("/admin/riwayat-notifikasi")
def admin_riwayat_notifikasi():
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    return render_template("06.RiwayatNotifikasi.html", riwayat=riwayat)

@app.route("/admin/riwayat-aktivitas")
def admin_riwayat_aktivitas():
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    daftar_admin = list(set(d['admin'] for d in data_aktivitas))
    return render_template("24.RiwayatAktivitas.html", daftar_admin=daftar_admin)

@app.route("/admin/api/aktivitas")
def admin_api_aktivitas():
    search = request.args.get('search', '').lower()
    admin_filter = request.args.get('admin', 'Pilihan Admin')
    tanggal = request.args.get('tanggal', '')
    hasil = data_aktivitas
    if search:
        hasil = [d for d in hasil if search in d['catatan'].lower()]
    if admin_filter != 'Pilihan Admin':
        hasil = [d for d in hasil if admin_filter in d['admin']]
    if tanggal:
        hasil = [d for d in hasil if d['waktu'].startswith(tanggal)]
    return jsonify(hasil)

@app.route("/admin/hapus-semua", methods=["POST"])
def admin_hapus_semua():
    global notifikasi, riwayat
    riwayat.extend(notifikasi)
    notifikasi.clear()
    return redirect("/admin/notifikasi")

@app.route('/admin/hapus-notif/<int:index>', methods=['POST'])
def admin_hapus_notif(index):
    if index < len(notifikasi):
        riwayat.append(notifikasi[index])
        notifikasi.pop(index)
    return redirect('/admin/notifikasi')

@app.route('/admin/hapus-riwayat-satuan/<int:index>', methods=['POST'])
def admin_hapus_riwayat_satuan(index):
    if index < len(riwayat):
        riwayat.pop(index)
    return redirect('/admin/riwayat')

@app.route("/admin/hapus-riwayat", methods=["POST"])
def admin_hapus_riwayat():
    global riwayat
    riwayat.clear()
    return redirect("/admin/riwayat")

# ============================================================
# ADMIN: DASHBOARD
# ============================================================

@app.route('/admin/dashboard')
def admin_dashboard():
    if not session.get('user'):
        return redirect(url_for('admin_login'))

    totalProduk = len(data_barang)
    stokMenipis = len([b for b in data_barang if b['stok'] <= 10])
    
    today_str = datetime.now().strftime("%Y-%m-%d")
    today_trx = [p for p in pesanan if p['tanggal'] == today_str]
    transaksiHariIni = len(today_trx) if today_trx else len(pesanan)
    
    total_income = sum(p['total'] for p in pesanan)
    pendapatan = formatRp(total_income)

    hasil_penjualan = {
        "Senin": {"makanan": 5, "minuman": 2, "alat tulis": 2},
        "Selasa": {"makanan": 0, "minuman": 5, "alat tulis": 5},
        "Rabu": {"makanan": 5, "minuman": 5, "alat tulis": 5},
        "Kamis": {"makanan": 6, "minuman": 7, "alat tulis": 5},
        "Jumat": {"makanan": 10, "minuman": 5, "alat tulis": 7},
    }

    y_hasilPenjualan = []
    for hari in hasil_penjualan:
        hasil = sum(hasil_penjualan[hari].values())
        y_hasilPenjualan.append(hasil)

    return render_template('07.Beranda_admin.html',
                           data_penjualan=y_hasilPenjualan,
                           card_product=totalProduk,
                           card_stock=stokMenipis,
                           card_transaction=transaksiHariIni,
                           card_income=pendapatan,
                           grafik_penjualan=hasil_penjualan)

# ============================================================
# ADMIN: KELOLA AKUN PENJUAL
# ============================================================

data_penjual = [
    {"id": 1, "nama": "Zulfikar Aril", "email": "aril_10@gmail.com", "status": "Aktif", "foto": "profile.png"},
    {"id": 2, "nama": "Nafilah Yasmin", "email": "yasmin18@gmail.com", "status": "Tidak aktif", "foto": "profile.png"},
    {"id": 3, "nama": "Febriyanto", "email": "febri1711@gmail.com", "status": "Aktif", "foto": "profile.png"},
    {"id": 4, "nama": "Shafira Amelia", "email": "shafiramel@gmail.com", "status": "Aktif", "foto": "profile.png"},
]

@app.route('/admin/kelola_akun_penjual')
def admin_kelola_akun_penjual():
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    
    keyword = request.args.get('cari', '').lower()
    if keyword:
        filtered = [p for p in data_penjual if keyword in p['nama'].lower() or keyword in p['email'].lower()]
    else:
        filtered = data_penjual
        
    return render_template('08.pengelola_akun_penjual.html', penjual=filtered, keyword=keyword)

@app.route('/admin/tambah-akun', methods=['GET', 'POST'])
def admin_tambah_akun():
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    if request.method == 'POST':
        max_id = max([p['id'] for p in data_penjual], default=0)
        akun_baru = {
            'id': max_id + 1,
            'nama': request.form['nama'],
            'email': request.form['email'],
            'status': request.form['status'],
            'foto': 'profile.png'
        }
        data_penjual.append(akun_baru)
        return redirect(url_for('admin_kelola_akun_penjual'))
    return render_template('08.tambah_akun.html')

@app.route('/admin/edit-akun/<int:id>', methods=['GET', 'POST'])
def admin_edit_akun(id):
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    
    akun = next((p for p in data_penjual if p['id'] == id), None)
    if not akun:
        return redirect(url_for('admin_kelola_akun_penjual'))
        
    if request.method == 'POST':
        akun['nama'] = request.form['nama']
        akun['email'] = request.form['email']
        akun['status'] = request.form['status']
        return redirect(url_for('admin_kelola_akun_penjual'))
    return render_template('08.edit_akun.html', akun=akun)

@app.route('/admin/hapus-akun/<int:id>')
def admin_hapus_akun(id):
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    
    global data_penjual
    data_penjual = [p for p in data_penjual if p['id'] != id]
    return redirect(url_for('admin_kelola_akun_penjual'))

# ============================================================
# ADMIN: MANAJEMEN BARANG
# ============================================================

@app.route('/admin/manajemen-barang')
def admin_manajemen_barang():
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    keyword = request.args.get('cari', '').lower()
    if keyword:
        filtered = [b for b in data_barang if keyword in b['nama'].lower() or keyword in b['kategori'].lower()]
    else:
        filtered = data_barang
    return render_template('09.manajemen_barang.html', data=filtered, keyword=request.args.get('cari', ''))

# ============================================================
# ADMIN: PENGISIAN BARANG
# ============================================================

@app.route('/admin/pengisian_barang')
def admin_pengisian_barang():
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    return render_template('10.pengisian_barang_.html')

@app.route('/admin/pengisian_barang/<int:id>')
def admin_pengisian_barang_restok(id):
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    barang = None
    for b in data_barang:
        if b['no'] == id:
            barang = b
            break
    if not barang:
        return redirect(url_for('admin_pengisian_barang'))
    return render_template('10.pengisian_barang_.html', barang=barang)

@app.route('/admin/tambah-data-barang')
def admin_tambah_data_barang():
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    return render_template('25.tambah_data_barang.html')

@app.route('/admin/simpan-barang-baru', methods=['POST'])
def admin_simpan_barang_baru():
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    kategori = request.form.get('kategori', '')
    nama_barang = request.form.get('nama_barang', '')
    harga_beli = int(request.form.get('harga_beli', 0))
    harga_jual = int(request.form.get('harga_jual', 0))
    jumlah = int(request.form.get('jumlah', 0))
    tanggal = request.form.get('tanggal', '')
    variasi = request.form.get('variasi', '')
    volume = request.form.get('volume', '')
    rasa = request.form.get('rasa', '')
    expired = request.form.get('expired', '')
    deskripsi = request.form.get('deskripsi', '')
    # Handle image upload
    gambar_path = ''
    file = request.files.get('gambar')
    if file and file.filename:
        filename = secure_filename(file.filename)
        upload_dir = app.config['UPLOAD_FOLDER']
        os.makedirs(upload_dir, exist_ok=True)
        file.save(os.path.join(upload_dir, filename))
        gambar_path = 'gambar/' + filename

    no_baru = max([b['no'] for b in data_barang], default=0) + 1
    data_barang.append({
        'no': no_baru, 'nama': nama_barang, 'berat': volume or '-',
        'stok': jumlah, 'harga': harga_jual, 'kategori': kategori,
        'tanggal': tanggal, 'gambar': gambar_path, 'rating': 0, 'emoji': '📦'
    })
    save_data_barang(data_barang)
    return render_template('17.-konfirmasi-barang.html',
        nama_barang=nama_barang, kategori=kategori,
        harga_beli=harga_beli, harga_jual=harga_jual, 
        jumlah=jumlah, tanggal=tanggal, variasi=variasi,
        ukuran=volume, rasa=rasa, expired=expired,
        deskripsi=deskripsi, gambar=gambar_path)

@app.route('/admin/simpan', methods=['POST'])
def admin_simpan():
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    kategori = request.form['kategori']
    nama_barang = request.form['nama_barang']
    tanggal = request.form.get('tanggal', '')
    jumlah = request.form['jumlah']
    harga_beli = request.form.get('harga_beli', '0')
    harga_jual = request.form.get('harga_jual', request.form.get('harga', '0'))
    catatan = request.form.get('catatan', '')
    restok_id = request.form.get('restok_id')

    # If restok, update existing item stock
    if restok_id:
        for b in data_barang:
            if str(b['no']) == str(restok_id):
                b['stok'] = b.get('stok', 0) + int(jumlah)
                save_data_barang(data_barang)
                break

    return render_template('10.rekap_barang.html', kategori=kategori, nama_barang=nama_barang, tanggal=tanggal, jumlah=jumlah, harga=harga_jual, catatan=catatan)

@app.route('/admin/konfirmasi-barang')
def admin_konfirmasi_barang():
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    nama_barang = request.args.get('nama_barang', '-')
    kategori = request.args.get('kategori', '-')
    variasi = request.args.get('variasi', '-')
    ukuran = request.args.get('ukuran', '-')
    rasa = request.args.get('rasa', '-')
    expired = request.args.get('expired', '-')
    deskripsi = request.args.get('deskripsi', '-')
    harga_beli = request.args.get('harga_beli', '0')
    harga_jual = request.args.get('harga_jual', '0')
    jumlah = request.args.get('jumlah', '0')
    tanggal = request.args.get('tanggal', '-')
    
    return render_template('17.-konfirmasi-barang.html', 
        nama_barang=nama_barang, kategori=kategori, variasi=variasi,
        ukuran=ukuran, rasa=rasa, expired=expired, deskripsi=deskripsi,
        harga_beli=harga_beli, harga_jual=harga_jual, jumlah=jumlah, 
        tanggal=tanggal)

# ============================================================
# ADMIN: STOK TERSEDIA
# ============================================================

@app.route('/admin/stok-tersedia')
def admin_stok_tersedia():
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    keyword = request.args.get('cari', '').lower()
    kategori = request.args.get('kategori', '')
    page = int(request.args.get('page', 1))
    per_page = 5

    filtered = data_barang
    if keyword:
        filtered = [b for b in filtered if keyword in b['nama'].lower()]
    if kategori:
        filtered = [b for b in filtered if b['kategori'] == kategori]

    total_halaman = max(1, (len(filtered) + per_page - 1) // per_page)
    start = (page - 1) * per_page
    data_page = filtered[start:start + per_page]

    return render_template('14.-stoktersedia.html',
        data=data_page, page=page, total_halaman=total_halaman,
        keyword=request.args.get('cari', ''), kategori=kategori)

@app.route('/admin/stok-tersedia/edit/<int:id>', methods=['GET', 'POST'])
def admin_stok_tersedia_edit(id):
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    barang = None
    for b in data_barang:
        if b['no'] == id:
            barang = b
            break
    if not barang:
        return "Barang tidak ditemukan", 404
    if request.method == 'POST':
        barang['nama'] = request.form.get('nama', barang['nama'])
        barang['berat'] = request.form.get('berat', barang['berat'])
        barang['kategori'] = request.form.get('kategori', barang['kategori'])
        barang['stok'] = int(request.form.get('stok', barang['stok']))
        barang['harga'] = int(request.form.get('harga', barang['harga']))
        barang['satuan'] = request.form.get('satuan', barang.get('satuan', ''))
        barang['tanggal_restok'] = request.form.get('tanggal_restok', barang.get('tanggal_restok', ''))
        barang['expired'] = request.form.get('expired', barang.get('expired', ''))
        barang['alasan'] = request.form.get('alasan', barang.get('alasan', ''))
        save_data_barang(data_barang)
        return redirect(url_for('admin_stok_tersedia'))
    return render_template('14.-stoktersedia_edit.html', barang=barang)
 
@app.route('/admin/stok-tersedia/hapus/<int:id>', methods=['POST'])
def admin_stok_tersedia_hapus(id):
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    global data_barang
    data_barang = [b for b in data_barang if b['no'] != id]
    save_data_barang(data_barang)
    return redirect(url_for('admin_stok_tersedia'))
 
# ============================================================
# ADMIN: CETAK LAPORAN
# ============================================================
 
data_barang = load_data_barang()

@app.route('/admin/cetak_laporan', methods=['GET', 'POST'])
def admin_cetak_laporan():
    if not session.get('user'):
        return redirect(url_for('admin_login'))

    tanggal_awal = request.args.get('tanggal_awal', '')
    tanggal_akhir = request.args.get('tanggal_akhir', '')

    data_transaksi = []
    for p in pesanan:
        if (not tanggal_awal or p['tanggal'] >= tanggal_awal) and (not tanggal_akhir or p['tanggal'] <= tanggal_akhir):
            total_barang = sum(b['harga'] * b['jumlah'] for b in p['barang'])
            data_transaksi.append({
                'tanggal': p['tanggal'],
                'id': p['id'],
                'jumlah_item': sum(b['jumlah'] for b in p['barang']),
                'total': total_barang
            })

    # Tambah data sample jika pesanan kosong dan no filters were set
    if not data_transaksi and not (tanggal_awal or tanggal_akhir):
        data_transaksi = [
            {'tanggal': '2026-05-01', 'id': 'TRX001', 'jumlah_item': 3, 'total': 15000},
            {'tanggal': '2026-05-02', 'id': 'TRX002', 'jumlah_item': 5, 'total': 25000},
            {'tanggal': '2026-05-03', 'id': 'TRX003', 'jumlah_item': 2, 'total': 10000},
        ]

    total_pendapatan = sum(t['total'] for t in data_transaksi)
    modal_barang = int(total_pendapatan * 0.7)
    untung_rugi = total_pendapatan - modal_barang

    return render_template('12.-cetaklaporan.html',
        barang=data_barang, formatRp=formatRp,
        total_pendapatan=total_pendapatan,
        modal_barang=modal_barang,
        untung_rugi=untung_rugi,
        data_transaksi=data_transaksi,
        data_barang_list=data_barang,
        pesanan_list=pesanan,
        tanggal_awal=tanggal_awal,
        tanggal_akhir=tanggal_akhir)

@app.route('/admin/cetak_transaksi_pdf')
def admin_cetak_transaksi_pdf():
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    tanggal_awal = request.args.get('tanggal_awal', '')
    tanggal_akhir = request.args.get('tanggal_akhir', '')
    filtered_pesanan = []
    for p in pesanan:
        if (not tanggal_awal or p['tanggal'] >= tanggal_awal) and (not tanggal_akhir or p['tanggal'] <= tanggal_akhir):
            filtered_pesanan.append(p)
    return render_template('12.-cetaklaporan_transaksi_pdf.html', pesanan=filtered_pesanan)

@app.route('/admin/cetak_transaksi_excel')
def admin_cetak_transaksi_excel():
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    tanggal_awal = request.args.get('tanggal_awal', '')
    tanggal_akhir = request.args.get('tanggal_akhir', '')
    filtered_pesanan = []
    for p in pesanan:
        if (not tanggal_awal or p['tanggal'] >= tanggal_awal) and (not tanggal_akhir or p['tanggal'] <= tanggal_akhir):
            filtered_pesanan.append(p)
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Transaksi"
    ws.append(['No', 'Tanggal', 'ID Transaksi', 'Nama Produk', 'Total', 'Metode', 'Status'])
    for i, p in enumerate(filtered_pesanan, 1):
        total_barang = sum(b['harga'] * b['jumlah'] for b in p['barang'])
        nama_produk = ', '.join(b['nama'] for b in p['barang'])
        tgl_fmt = format_kbbi_date(p['tanggal'])
        ws.append([i, tgl_fmt, p['id'], nama_produk, formatRp(total_barang), p['metode'], p['status']])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment;filename=laporan_transaksi.xlsx'}
    )

@app.route('/admin/pengaturan-laporan')
def admin_pengaturan_laporan():
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    tanggal_awal = request.args.get('tanggal_awal', '')
    tanggal_akhir = request.args.get('tanggal_akhir', '')
    laporan_type = request.args.get('laporan', 'barang')
    return render_template("16.pengaturan_laporan.html",
                           barang=data_barang,
                           tanggal_awal=tanggal_awal,
                           tanggal_akhir=tanggal_akhir,
                           laporan_type=laporan_type)

@app.route('/admin/cetak_laporan_barang')
def admin_cetak_laporan_barang():
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    total_nilai = sum(b['harga'] * b['stok'] for b in data_barang)
    return render_template('12.-cetaklaporan_barang.html',
        data_barang=data_barang,
        total_nilai=total_nilai)

@app.route('/admin/cetak_barang_pdf')
def admin_cetak_barang_pdf():
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    tanggal_awal = request.args.get('tanggal_awal', '')
    tanggal_akhir = request.args.get('tanggal_akhir', '')
    filtered_barang = []
    for b in data_barang:
        if (not tanggal_awal or b['tanggal'] >= tanggal_awal) and (not tanggal_akhir or b['tanggal'] <= tanggal_akhir):
            filtered_barang.append(b)
    total_nilai = sum(b['harga'] * b['stok'] for b in filtered_barang)
    return render_template('12.-cetaklaporan_barang_pdf.html',
        data_barang=filtered_barang,
        total_nilai=total_nilai)

@app.route('/admin/cetak_barang_excel')
def admin_cetak_barang_excel():
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    tanggal_awal = request.args.get('tanggal_awal', '')
    tanggal_akhir = request.args.get('tanggal_akhir', '')
    filtered_barang = []
    for b in data_barang:
        if (not tanggal_awal or b['tanggal'] >= tanggal_awal) and (not tanggal_akhir or b['tanggal'] <= tanggal_akhir):
            filtered_barang.append(b)
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Data Barang"
    ws.append(['No', 'Nama', 'Stok', 'Harga', 'Kategori', 'Tanggal', 'Total'])
    for b in filtered_barang:
        ws.append([b['no'], b['nama'], b['stok'], formatRp(b['harga']), b['kategori'], format_kbbi_date(b['tanggal']), formatRp(b['harga'] * b['stok'])])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment;filename=laporan_data_barang.xlsx'}
    )

@app.route('/admin/laporan_penjualan', methods=['GET', 'POST'])
def admin_laporan_penjualan():
    if not session.get('user'):
        return redirect(url_for('admin_login'))

    # Kalender
    tanggal_awal = request.values.get("tanggal_awal") or "2026-01-01"
    tanggal_akhir = request.values.get("tanggal_akhir") or "2026-12-31"
    if tanggal_akhir < tanggal_awal:
        tanggal_akhir = tanggal_awal

    def Rupiah(angka):
        return "Rp{:,.2f}".format(angka).replace(",", "X").replace(".", ",").replace("X", ".")

    # Data harian
    data_harian = {
        "2026-04-06": {"hari": "Senin", "totalTransaksi": 35, "totalPendapatan": 98000, "modalBarang": 74000},
        "2026-04-07": {"hari": "Selasa", "totalTransaksi": 38, "totalPendapatan": 110000, "modalBarang": 83000},
        "2026-04-08": {"hari": "Rabu", "totalTransaksi": 48, "totalPendapatan": 200000, "modalBarang": 150000},
        "2026-04-09": {"hari": "Kamis", "totalTransaksi": 37, "totalPendapatan": 104000, "modalBarang": 78000},
        "2026-04-10": {"hari": "Jumat", "totalTransaksi": 39, "totalPendapatan": 110000, "modalBarang": 83000},
        "2026-05-01": {"hari": "Jumat", "totalTransaksi": 35, "totalPendapatan": 100000, "modalBarang": 75000},
        "2026-05-04": {"hari": "Senin", "totalTransaksi": 35, "totalPendapatan": 100000, "modalBarang": 75000},
        "2026-05-05": {"hari": "Selasa", "totalTransaksi": 35, "totalPendapatan": 100000, "modalBarang": 75000},
        "2026-05-06": {"hari": "Rabu", "totalTransaksi": 34, "totalPendapatan": 98000, "modalBarang": 74000},
        "2026-05-07": {"hari": "Kamis", "totalTransaksi": 44, "totalPendapatan": 118000, "modalBarang": 89000},
        "2026-05-08": {"hari": "Jumat", "totalTransaksi": 34, "totalPendapatan": 98000, "modalBarang": 74000},
        "2026-05-11": {"hari": "Senin", "totalTransaksi": 33, "totalPendapatan": 94000, "modalBarang": 71000},
        "2026-05-12": {"hari": "Selasa", "totalTransaksi": 35, "totalPendapatan": 101000, "modalBarang": 76000},
        "2026-05-13": {"hari": "Rabu", "totalTransaksi": 33, "totalPendapatan": 95000, "modalBarang": 71000},
        "2026-05-14": {"hari": "Kamis", "totalTransaksi": 34, "totalPendapatan": 98000, "modalBarang": 74000},
        "2026-05-15": {"hari": "Jumat", "totalTransaksi": 33, "totalPendapatan": 96000, "modalBarang": 72000},
    }

    # 1. Filter real transactions from pesanan
    real_transactions = []
    for p in pesanan:
        if tanggal_awal <= p['tanggal'] <= tanggal_akhir:
            real_transactions.append(p)

    if real_transactions:
        totalTransaksi = len(real_transactions)
        totalPendapatan_num = sum(p['total'] for p in real_transactions)
        modalBarang_num = int(totalPendapatan_num * 0.75)
        untungRugi_num = totalPendapatan_num - modalBarang_num

        # Diagram produk terlaris
        total_produk = {}
        for p in real_transactions:
            for b in p.get('barang', []):
                nama = b['nama']
                jumlah = b['jumlah']
                total_produk[nama] = total_produk.get(nama, 0) + jumlah
        namaProduk = list(total_produk.keys())
        jumlahProduk = list(total_produk.values())
        
        warna_produk = {
            "air mineral": "blue", "roti aoka": "orange", "pensil": "red", 
            "pulpen": "yellow", "buku tulis": "purple", "pop mie": "green", 
            "penghapus": "pink", "bolpoin": "cyan", "teh botol": "teal", "ultra milk": "brown"
        }
        warnadiagramPT = []
        for prod in namaProduk:
            prod_lower = prod.lower()
            if prod_lower in warna_produk:
                warnadiagramPT.append(warna_produk[prod_lower])
            else:
                warnadiagramPT.append(random.choice(["blue", "orange", "red", "yellow", "purple", "green", "pink", "cyan", "teal", "brown"]))

        # Diagram pendapatan perkategori
        prod_categories = {b['nama']: b['kategori'] for b in data_barang}
        total_kategori = {}
        for p in real_transactions:
            for b in p.get('barang', []):
                nama = b['nama']
                revenue = b['harga'] * b['jumlah']
                cat = prod_categories.get(nama, "Lainnya")
                total_kategori[cat] = total_kategori.get(cat, 0) + revenue
                
        kategori = list(total_kategori.keys())
        pendapatanPerkategori = list(total_kategori.values())
        warna_kategori = {"minuman": "blue", "alat tulis": "red", "makanan": "green", "lainnya": "gray"}
        warnadiagramPP = [warna_kategori.get(k.lower(), "gray") for k in kategori]

        # Diagram perbandingan pendapatan tiap bulan
        try:
            target_year = int(tanggal_awal.split('-')[0])
        except:
            target_year = 2026
            
        pendapatanbulan = ["Januari","Februari","Maret","April","Mei","Juni","Juli","Agustus","September","Oktober","November","Desember"]
        pendapatanPerbulan = []
        for bulan in range(1, 13):
            total = 0
            hasData = False
            for p in pesanan:
                try:
                    pyear, pmonth, pday = map(int, p['tanggal'].split('-'))
                    if pyear == target_year and pmonth == bulan:
                        if tanggal_awal <= p['tanggal'] <= tanggal_akhir:
                            total += p['total']
                            hasData = True
                except:
                    continue
            pendapatanPerbulan.append(total if hasData else None)

        # Diagram perbandingan jumlah transaksi tiap bulan
        transaksiperbulan = ["Januari","Februari","Maret","April","Mei","Juni","Juli","Agustus","September","Oktober","November","Desember"]
        jumlahtransaksiPerbulan = []
        for bulan in range(1, 13):
            total = 0
            hasData = False
            for p in pesanan:
                try:
                    pyear, pmonth, pday = map(int, p['tanggal'].split('-'))
                    if pyear == target_year and pmonth == bulan:
                        if tanggal_awal <= p['tanggal'] <= tanggal_akhir:
                            total += 1
                            hasData = True
                except:
                    continue
            jumlahtransaksiPerbulan.append(total if hasData else None)
    else:
        # Data diagram produk terlaris
        data_chart_produk_terlaris_perhari = {
            "2026-04-06": {"namaProduk": ["Air mineral", "Aoka", "Pensil", "Pulpen", "Buku"], "jumlahProduk": [10, 5, 10, 7, 3], "warnadiagramPT": ["Blue", "Orange", "Red", "Yellow", "Purple"]},
            "2026-04-07": {"namaProduk": ["Air mineral", "Aoka", "Pensil", "Pulpen", "Buku"], "jumlahProduk": [12, 6, 8, 8, 4], "warnadiagramPT": ["Blue", "Orange", "Red", "Yellow", "Purple"]},
            "2026-04-08": {"namaProduk": ["Air mineral","Aoka","Pensil","Pulpen","Buku","Pop mie"], "jumlahProduk": [11, 6, 9, 7, 3, 12], "warnadiagramPT": ["Blue", "Orange", "Red", "Yellow", "Purple", "Green"]},
            "2026-04-09": {"namaProduk": ["Air mineral", "Aoka", "Pensil", "Pulpen", "Buku"], "jumlahProduk": [10, 7, 10, 7, 3], "warnadiagramPT": ["Blue", "Orange", "Red", "Yellow", "Purple"]},
            "2026-04-10": {"namaProduk": ["Air mineral", "Aoka", "Pensil", "Pulpen", "Buku"], "jumlahProduk": [12, 6, 9, 8, 4], "warnadiagramPT": ["Blue", "Orange", "Red", "Yellow", "Purple"]},
            "2026-05-01": {"namaProduk": ["Air mineral", "Aoka", "Pensil", "Pulpen", "Buku"], "jumlahProduk": [7, 8, 9, 7, 4], "warnadiagramPT": ["Blue", "Orange", "Red", "Yellow", "Purple"]},
            "2026-05-04": {"namaProduk": ["Air mineral", "Aoka", "Pensil", "Pulpen", "Buku"], "jumlahProduk": [7, 8, 9, 7, 4], "warnadiagramPT": ["Blue", "Orange", "Red", "Yellow", "Purple"]},
            "2026-05-05": {"namaProduk": ["Air mineral", "Aoka", "Pensil", "Pulpen", "Buku"], "jumlahProduk": [7, 8, 9, 7, 4], "warnadiagramPT": ["Blue", "Orange", "Red", "Yellow", "Purple"]},
            "2026-05-06": {"namaProduk": ["Air mineral", "Aoka", "Pensil", "Pulpen", "Buku"], "jumlahProduk": [7, 8, 8, 7, 4], "warnadiagramPT": ["Blue", "Orange", "Red", "Yellow", "Purple"]},
            "2026-05-07": {"namaProduk": ["Air mineral","Aoka","Pensil","Pulpen","Buku","Penghapus"], "jumlahProduk": [7, 8, 8, 7, 4, 10], "warnadiagramPT": ["Blue", "Orange", "Red", "Yellow", "Purple", "Pink"]},
            "2026-05-08": {"namaProduk": ["Air mineral", "Aoka", "Pensil", "Pulpen", "Buku"], "jumlahProduk": [7, 8, 8, 7, 4], "warnadiagramPT": ["Blue", "Orange", "Red", "Yellow", "Purple"]},
            "2026-05-11": {"namaProduk": ["Air mineral", "Aoka", "Pensil", "Pulpen", "Buku"], "jumlahProduk": [7, 8, 8, 7, 3], "warnadiagramPT": ["Blue", "Orange", "Red", "Yellow", "Purple"]},
            "2026-05-12": {"namaProduk": ["Air mineral", "Aoka", "Pensil", "Pulpen", "Buku"], "jumlahProduk": [8, 7, 9, 7, 4], "warnadiagramPT": ["Blue", "Orange", "Red", "Yellow", "Purple"]},
            "2026-05-13": {"namaProduk": ["Air mineral", "Aoka", "Pensil", "Pulpen", "Buku"], "jumlahProduk": [7, 7, 8, 8, 3], "warnadiagramPT": ["Blue", "Orange", "Red", "Yellow", "Purple"]},
            "2026-05-14": {"namaProduk": ["Air mineral", "Aoka", "Pensil", "Pulpen", "Buku"], "jumlahProduk": [7, 8, 9, 7, 3], "warnadiagramPT": ["Blue", "Orange", "Red", "Yellow", "Purple"]},
            "2026-05-15": {"namaProduk": ["Air mineral", "Aoka", "Pensil", "Pulpen", "Buku"], "jumlahProduk": [8, 7, 8, 7, 3], "warnadiagramPT": ["Blue", "Orange", "Red", "Yellow", "Purple"]},
        }

        # Data Diagram pendapatan perkategori
        data_chart_pendapatan_perkategori_perhari = {
            "2026-04-06": {"kategori": ["Minuman", "Alat tulis", "Makanan"], "pendapatanPerkategori": [30000, 53000, 15000], "warnadiagramPP": ["Blue", "Red", "Green"]},
            "2026-04-07": {"kategori": ["Minuman", "Alat tulis", "Makanan"], "pendapatanPerkategori": [36000, 56000, 18000], "warnadiagramPP": ["Blue", "Red", "Green"]},
            "2026-04-08": {"kategori": ["Minuman", "Alat tulis", "Makanan"], "pendapatanPerkategori": [33000, 53000, 114000], "warnadiagramPP": ["Blue", "Red", "Green"]},
            "2026-04-09": {"kategori": ["Minuman", "Alat tulis", "Makanan"], "pendapatanPerkategori": [30000, 53000, 21000], "warnadiagramPP": ["Blue", "Red", "Green"]},
            "2026-04-10": {"kategori": ["Minuman", "Alat tulis", "Makanan"], "pendapatanPerkategori": [36000, 56000, 18000], "warnadiagramPP": ["Blue", "Red", "Green"]},
            "2026-05-01": {"kategori": ["Minuman", "Alat tulis", "Makanan"], "pendapatanPerkategori": [21000, 55000, 24000], "warnadiagramPP": ["Blue", "Red", "Green"]},
            "2026-05-04": {"kategori": ["Minuman", "Alat tulis", "Makanan"], "pendapatanPerkategori": [21000, 55000, 24000], "warnadiagramPP": ["Blue", "Red", "Green"]},
            "2026-05-05": {"kategori": ["Minuman", "Alat tulis", "Makanan"], "pendapatanPerkategori": [21000, 55000, 24000], "warnadiagramPP": ["Blue", "Red", "Green"]},
            "2026-05-06": {"kategori": ["Minuman", "Alat tulis", "Makanan"], "pendapatanPerkategori": [21000, 53000, 24000], "warnadiagramPP": ["Blue", "Red", "Green"]},
            "2026-05-07": {"kategori": ["Minuman", "Alat tulis", "Makanan"], "pendapatanPerkategori": [21000, 73000, 24000], "warnadiagramPP": ["Blue", "Red", "Green"]},
            "2026-05-08": {"kategori": ["Minuman", "Alat tulis", "Makanan"], "pendapatanPerkategori": [21000, 53000, 24000], "warnadiagramPP": ["Blue", "Red", "Green"]},
            "2026-05-11": {"kategori": ["Minuman", "Alat tulis", "Makanan"], "pendapatanPerkategori": [21000, 49000, 24000], "warnadiagramPP": ["Blue", "Red", "Green"]},
            "2026-05-12": {"kategori": ["Minuman", "Alat tulis", "Makanan"], "pendapatanPerkategori": [24000, 56000, 21000], "warnadiagramPP": ["Blue", "Red", "Green"]},
            "2026-05-13": {"kategori": ["Minuman", "Alat tulis", "Makanan"], "pendapatanPerkategori": [21000, 53000, 21000], "warnadiagramPP": ["Blue", "Red", "Green"]},
            "2026-05-14": {"kategori": ["Minuman", "Alat tulis", "Makanan"], "pendapatanPerkategori": [21000, 53000, 24000], "warnadiagramPP": ["Blue", "Red", "Green"]},
            "2026-05-15": {"kategori": ["Minuman", "Alat tulis", "Makanan"], "pendapatanPerkategori": [24000, 51000, 21000], "warnadiagramPP": ["Blue", "Red", "Green"]},
        }

        # kolom informasi penting
        totalTransaksi = 0
        totalPendapatan_num = 0
        modalBarang_num = 0
        for tanggal, data in data_harian.items():
            if tanggal_awal <= tanggal <= tanggal_akhir:
                totalTransaksi += data["totalTransaksi"]
                totalPendapatan_num += data["totalPendapatan"]
                modalBarang_num += data["modalBarang"]
        untungRugi_num = totalPendapatan_num - modalBarang_num

        # Diagram produk terlaris
        total_produk = {}
        for tanggal, chart in data_chart_produk_terlaris_perhari.items():
            if tanggal_awal <= tanggal <= tanggal_akhir:
                for i in range(len(chart["namaProduk"])):
                    nama = chart["namaProduk"][i]
                    jumlah = chart["jumlahProduk"][i]
                    if nama not in total_produk:
                        total_produk[nama] = 0
                    total_produk[nama] += jumlah
        namaProduk = list(total_produk.keys())
        jumlahProduk = list(total_produk.values())
        warna_produk = {"Air mineral": "blue", "Aoka": "orange", "Pensil": "red", "Pulpen": "yellow", "Buku": "purple", "Pop mie": "green", "Penghapus": "pink"}
        warnadiagramPT = [warna_produk.get(produk, "gray") for produk in namaProduk]

        # Diagram pendapatan perkategori
        total_kategori = {}
        for tanggal, chart in data_chart_pendapatan_perkategori_perhari.items():
            if tanggal_awal <= tanggal <= tanggal_akhir:
                for i in range(len(chart["kategori"])):
                    nama = chart["kategori"][i]
                    jumlah = chart["pendapatanPerkategori"][i]
                    if nama not in total_kategori:
                        total_kategori[nama] = 0
                    total_kategori[nama] += jumlah
        kategori = list(total_kategori.keys())
        pendapatanPerkategori = list(total_kategori.values())
        warna_kategori = {"Minuman": "blue", "Alat tulis": "red", "Makanan": "green"}
        warnadiagramPP = [warna_kategori.get(k, "gray") for k in kategori]

        # Diagram perbandingan pendapatan tiap bulan
        pendapatanbulan = ["Januari","Februari","Maret","April","Mei","Juni","Juli","Agustus","September","Oktober","November","Desember"]
        pendapatanPerbulan = []
        for bulan in range(1, 13):
            total = 0
            hasData = False
            for tanggal, data in data_harian.items():
                tahun, bln, hari = map(int, tanggal.split("-"))
                if tanggal_awal <= tanggal <= tanggal_akhir and bln == bulan:
                    total += data["totalPendapatan"]
                    hasData = True
            pendapatanPerbulan.append(total if hasData else None)

        # Diagram perbandingan jumlah transaksi tiap bulan
        transaksiperbulan = ["Januari","Februari","Maret","April","Mei","Juni","Juli","Agustus","September","Oktober","November","Desember"]
        jumlahtransaksiPerbulan = []
        for bulan in range(1, 13):
            total = 0
            hasData = False
            for tanggal, data in data_harian.items():
                tahun, bln, hari = map(int, tanggal.split("-"))
                if tanggal_awal <= tanggal <= tanggal_akhir and bln == bulan:
                    total += data["totalTransaksi"]
                    hasData = True
            jumlahtransaksiPerbulan.append(total if hasData else None)

    # Format numbers to Rupiah string
    totalPendapatan = Rupiah(totalPendapatan_num)
    modalBarang = Rupiah(modalBarang_num)
    if untungRugi_num >= 0:
        untungRugi = "+" + Rupiah(untungRugi_num)
    else:
        untungRugi = "-" + Rupiah(abs(untungRugi_num))

    return render_template('15.Laporan_Penjualan.html',
        tanggal_awal=tanggal_awal, tanggal_akhir=tanggal_akhir,
        TT=totalTransaksi, TP=totalPendapatan, MB=modalBarang, UG=untungRugi,
        NP=namaProduk, JP=jumlahProduk, WPT=warnadiagramPT,
        KT=kategori, PPK=pendapatanPerkategori, WPP=warnadiagramPP,
        PL=pendapatanbulan, PPB=pendapatanPerbulan,
        TSP=transaksiperbulan, JTP=jumlahtransaksiPerbulan)

# ============================================================
# ADMIN: SIAPKAN PESANAN
# ============================================================

status_tunai = ["Disiapkan", "Siap diambil", "Menunggu Pembayaran", "Sudah diambil"]
status_qris  = ["Disiapkan", "Siap diambil", "Sudah diambil"]

pesanan = load_pesanan()

def formatRp(rupiah):
    try:
        rupiah = float(rupiah)
    except (ValueError, TypeError):
        return rupiah
    return "Rp {:,.0f}".format(rupiah).replace(",", ".") + ",00"

app.jinja_env.globals.update(formatRp=formatRp)

def get_status_list(pesan):
    return status_qris if pesan['metode'].lower() == 'qris' else status_tunai

@app.route('/admin/siapkan-pesanan', methods=['GET', 'POST'])
def admin_siapkan_pesanan():
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    if request.method == 'POST':
        trx_id = request.form["trx_id"]
        new_status = request.form.get("status")
        for i in pesanan:
            if i['id'] == trx_id:
                if new_status:
                    ganti = get_status_list(i)
                    if new_status in ganti:
                        i['status'] = new_status
                else:
                    ganti = get_status_list(i)
                    ubah = ganti.index(i['status'])
                    if ubah + 1 < len(ganti):
                        i['status'] = ganti[ubah + 1]
                break
        save_pesanan(pesanan)
        return redirect('/admin/siapkan-pesanan')
    # Pagination
    page = request.args.get('page', 1, type=int)
    per_page = 20
    total = len(pesanan)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    start = (page - 1) * per_page
    end = start + per_page
    pesanan_page = pesanan[start:end]
    bulan_id = ['Januari','Februari','Maret','April','Mei','Juni','Juli','Agustus','September','Oktober','November','Desember']
    for p in pesanan_page:
        try:
            from datetime import datetime
            tgl = datetime.strptime(p['tanggal'], '%Y-%m-%d')
            p['tanggal_fmt'] = f"{tgl.day:02d}/{tgl.month:02d}/{tgl.year}"
        except:
            p['tanggal_fmt'] = p['tanggal']
    return render_template('19.SiapkanPesanan.html', pesanan=pesanan_page, formatRp=formatRp,
                           status=get_status_list, page=page, total_pages=total_pages, total=total)

@app.route("/admin/hapus-barang", methods=["POST"])
def admin_hapus_barang():
    trx_id = request.form["trx_id"]
    nama_barang = request.form["nama_barang"]
    for p in pesanan:
        if p["id"] == trx_id:
            dihapus = None
            for b in p["barang"]:
                if b["nama"] == nama_barang:
                    dihapus = b
                    break
            if dihapus:
                dihapus["jumlah"] -= 1
                # Restore stock in data_barang
                for db_item in data_barang:
                    if db_item['nama'] == nama_barang:
                        db_item['stok'] += 1
                        break
                save_data_barang(data_barang)
                
                if dihapus["jumlah"] <= 0:
                    p["barang"].remove(dihapus)
                p["total"] = hitung_total_barang(p["barang"])
                p["refund"] = p["total_awal"] - p["total"]
                save_pesanan(pesanan)
            break
    return redirect("/admin/siapkan-pesanan")

# ============================================================
# ADMIN: CEK PEMBAYARAN
# ============================================================

@app.route('/admin/cek-pembayaran')
def admin_cek_pembayaran():
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    cari = request.args.get('cari', '').strip().lower()
    page = request.args.get('page', 1, type=int)
    per_page = 20
    if cari:
        hasil = [p for p in pesanan if cari in p['id'].lower() or cari in p['pelanggan'].lower()]
    else:
        hasil = list(pesanan)
    total = len(hasil)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    start = (page - 1) * per_page
    end = start + per_page
    transaksi = hasil[start:end]
    return render_template('21.cek_pembayaran.html', pesanan=transaksi,
                           page=page, total_pages=total_pages, total=total,
                           keyword=cari)

@app.route('/admin/cek-pembayaran/detail/<trx_id>')
def admin_cek_pembayaran_detail(trx_id):
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    order = None
    for p in pesanan:
        if p['id'] == trx_id:
            order = p
            break
    if not order:
        return "Transaksi tidak ditemukan", 404
    return render_template('21.cek_pembayaran_detail.html', order=order)

# ============================================================
# ADMIN: PENGATURAN
# ============================================================

app.config['UPLOAD_FOLDER'] = os.path.join(BASE_DIR, 'static', 'gambar')
DATA_FILE = os.path.join(BASE_DIR, 'data_koperasi.json')

def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r') as f:
            return json.load(f)
    return {
        'nama': 'Koperasi Sekolah',
        'deskripsi': 'Koperasi untuk siswa dan guru',
        'alamat': 'Jl. Pendidikan No. 1',
        'telepon': '0341-123456',
        'jam': '07.00 - 15.00',
        'hari': 'Senin - Jumat',
        'logo': 'image/logo_3.png'
    }

def save_data(data):
    with open(DATA_FILE, 'w') as f:
        json.dump(data, f)

data_koperasi = load_data()

class _NamaKoperasi:
    __slots__ = ()
    def __str__(self):
        return data_koperasi.get('nama', 'Koperasi Sekolah')
    def __html__(self):
        return data_koperasi.get('nama', 'Koperasi Sekolah')

app.jinja_env.globals['nama_koperasi'] = _NamaKoperasi()

@app.route('/admin/pengaturan', methods=['GET'])
def admin_pengaturan():
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    return render_template('22.pengaturan_umum.html', **data_koperasi)

@app.route('/admin/simpan_pengaturan', methods=['POST'])
def admin_simpan_pengaturan():
    if not session.get('user'):
        return redirect(url_for('admin_login'))
    global data_koperasi
    data_koperasi['nama'] = request.form.get('nama', data_koperasi['nama'])
    data_koperasi['deskripsi'] = request.form.get('deskripsi', data_koperasi['deskripsi'])
    data_koperasi['alamat'] = request.form.get('alamat', data_koperasi['alamat'])
    data_koperasi['telepon'] = request.form.get('telepon', data_koperasi['telepon'])
    data_koperasi['jam'] = request.form.get('jam', data_koperasi['jam'])
    data_koperasi['hari'] = request.form.get('hari', data_koperasi['hari'])

    # Handle logo file upload (from modal form field "gambar")
    file = request.files.get('gambar')
    if file and file.filename:
        filename = secure_filename(file.filename)
        upload_dir = app.config['UPLOAD_FOLDER']
        os.makedirs(upload_dir, exist_ok=True)
        file.save(os.path.join(upload_dir, filename))
        data_koperasi['logo'] = 'gambar/' + filename
        
        # Overwrite standard logo_3.png to globally update the entire app
        file.seek(0)
        file.save(os.path.join(BASE_DIR, 'static', 'image', 'logo_3.png'))

    save_data(data_koperasi)
    return redirect(url_for('admin_pengaturan'))

# ============================================================
# PEMBELI: LOGIN & LOGOUT
# ============================================================

@app.route('/pembeli/login', methods=['GET', 'POST'])
def pembeli_login():
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        if USERS.get(username) == password:
            session['user'] = username
            session['role'] = 'pembeli'
            session['nama'] = username
            return redirect(url_for('pembeli_home'))
        error = 'Nama akun atau kata sandi tidak cocok.'
    return render_template('login_pembeli.html', error=error)

@app.route('/pembeli/register', methods=['GET', 'POST'])
def pembeli_register():
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        if not username or not password:
            error = 'Nama akun dan kata sandi wajib diisi.'
        elif username in USERS:
            error = 'Nama akun sudah digunakan.'
        else:
            USERS[username] = password
            return redirect(url_for('pembeli_login'))
    return render_template('register_pembeli.html', error=error)

@app.route('/pembeli/logout')
def pembeli_logout():
    session.clear()
    return redirect(url_for('pembeli_login'))

# ============================================================
# PEMBELI: RESET PASSWORD (OTP FLOW)
# ============================================================

otp_storage = {}  # email -> otp code

@app.route('/pembeli/reset-password', methods=['GET', 'POST'])
def pembeli_reset_password():
    error = None
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        if not email:
            error = 'Email wajib diisi.'
        else:
            otp = str(random.randint(10000, 99999))
            otp_storage[email] = otp
            session['reset_email'] = email
            return redirect(url_for('pembeli_verifikasi_email'))
    return render_template('reset_pembeli.html', error=error)

@app.route('/pembeli/kirim-otp', methods=['GET', 'POST'])
def pembeli_kirim_otp():
    error = None
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        if not email:
            error = 'Email wajib diisi.'
        else:
            otp = str(random.randint(10000, 99999))
            otp_storage[email] = otp
            session['reset_email'] = email
            return redirect(url_for('pembeli_verifikasi_email'))
    return render_template('reset_pembeli.html', error=error)

@app.route('/pembeli/verifikasi-email', methods=['GET', 'POST'])
def pembeli_verifikasi_email():
    error = None
    email = session.get('reset_email', '')
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        session['reset_email'] = email
        return redirect(url_for('pembeli_verifikasi_otp'))
    return render_template('verifikasi_email_pembeli.html', error=error, email=email)

@app.route('/pembeli/verifikasi-otp', methods=['GET', 'POST'])
def pembeli_verifikasi_otp():
    error = None
    email = session.get('reset_email', '')
    if request.method == 'POST':
        kode = ''.join([
            request.form.get('kode1', ''),
            request.form.get('kode2', ''),
            request.form.get('kode3', ''),
            request.form.get('kode4', ''),
            request.form.get('kode5', ''),
        ])
        stored_otp = otp_storage.get(email, '')
        if kode == stored_otp:
            session['otp_verified'] = True
            return redirect(url_for('pembeli_ganti_password'))
        error = 'Kode OTP tidak cocok.'
    return render_template('verifikasi_email_pembeli.html', error=error, email=email)

@app.route('/pembeli/ganti-password', methods=['GET', 'POST'])
def pembeli_ganti_password():
    error = None
    if not session.get('otp_verified'):
        return redirect(url_for('pembeli_reset_password'))
    if request.method == 'POST':
        password = request.form.get('password', '').strip()
        confirm = request.form.get('confirm_password', '').strip()
        if not password or len(password) < 8:
            error = 'Kata sandi minimal 8 karakter.'
        elif password != confirm:
            error = 'Konfirmasi kata sandi tidak cocok.'
        else:
            email = session.get('reset_email', '')
            otp_storage.pop(email, None)
            session.pop('otp_verified', None)
            session.pop('reset_email', None)
            return render_template('notifikasi_berhasil_pembeli.html', message='Kata sandi berhasil diubah. Silakan login dengan kata sandi baru.')
    return render_template('reset_pembeli.html', error=error, show_password_form=True)

# ============================================================
# PEMBELI: HOME
# ============================================================

@app.route('/pembeli')
def pembeli_home():
    if not session.get('user'):
        return redirect(url_for('pembeli_login'))
    cart_count = sum(item['jumlah'] for item in cart.values())
    sorted_barang = sorted(data_barang, key=lambda b: (b['stok'] == 0, b['nama']))
    return render_template('14-kategorialattulis.html', barang=sorted_barang, cart_count=cart_count)

# ============================================================
# PEMBELI: KATEGORI ALAT TULIS
# ============================================================

@app.route('/pembeli/detail-barang/<int:id>')
def pembeli_detail_barang(id):
    if not session.get('user'):
        return redirect(url_for('pembeli_login'))
    barang = None
    for b in data_barang:
        if b['no'] == id:
            barang = {'id': b['no'], 'nama': b['nama'], 'harga': b['harga'], 'stok': b['stok'],
                      'gambar': b.get('gambar', ''), 'kategori': b['kategori'], 'rating': b.get('rating', 4),
                      'berat': b.get('berat', ''), 'deskripsi': f"{b['nama']} dari kategori {b['kategori']}."}
            break
    if not barang:
        return "Barang tidak ditemukan", 404
    return render_template('7-detailbarang.html', barang=barang, formatRp=formatRp)

# ============================================================
# PEMBELI: KERANJANG
# ============================================================

cart = {}

@app.route('/pembeli/tambah-keranjang', methods=['POST'])
def pembeli_tambah_keranjang():
    nama = request.form.get('nama', '').strip()
    harga = int(request.form.get('harga', 0))
    jumlah = int(request.form.get('jumlah', 1))
    gambar = request.form.get('gambar', '')
    
    barang = next((b for b in data_barang if b['nama'] == nama), None)
    stok_tersedia = barang['stok'] if barang else 0

    if nama and jumlah > 0:
        current_qty = cart.get(nama, {}).get('jumlah', 0)
        if current_qty + jumlah <= stok_tersedia:
            if nama in cart:
                cart[nama]['jumlah'] += jumlah
            else:
                cart[nama] = {'harga': harga, 'jumlah': jumlah, 'gambar': gambar}
        else:
            # Cap at available stock
            if stok_tersedia > 0:
                cart[nama] = {'harga': harga, 'jumlah': stok_tersedia, 'gambar': gambar}
    return redirect('/pembeli?added=1')

@app.route('/pembeli/update-keranjang', methods=['POST'])
def pembeli_update_keranjang():
    nama = request.form.get('nama', '').strip()
    aksi = request.form.get('aksi', '')
    if nama in cart:
        if aksi == 'tambah':
            barang = next((b for b in data_barang if b['nama'] == nama), None)
            stok_tersedia = barang['stok'] if barang else 0
            if cart[nama]['jumlah'] < stok_tersedia:
                cart[nama]['jumlah'] += 1
        elif aksi == 'kurang':
            cart[nama]['jumlah'] -= 1
            if cart[nama]['jumlah'] <= 0:
                del cart[nama]
        elif aksi == 'hapus':
            del cart[nama]
    return redirect('/pembeli/keranjang')

@app.route('/pembeli/keranjang')
def pembeli_keranjang():
    # Handle ?tambah=ID from detail barang page
    tambah_id = request.args.get('tambah')
    qty = int(request.args.get('qty', 1))
    if tambah_id:
        tambah_id = int(tambah_id)
        barang = None
        for b in data_barang:
            if b['no'] == tambah_id:
                barang = {'id': b['no'], 'nama': b['nama'], 'harga': b['harga'], 'stok': b['stok'],
                          'gambar': b.get('gambar', ''), 'kategori': b['kategori']}
                break
        
        if barang:
            stok_tersedia = barang.get('stok', 0)
            nama = barang['nama']
            gambar = barang.get('gambar', '')
            
            if stok_tersedia > 0:
                current_qty = cart.get(nama, {}).get('jumlah', 0)
                if current_qty + qty <= stok_tersedia:
                    if nama in cart:
                        cart[nama]['jumlah'] += qty
                    else:
                        cart[nama] = {'harga': barang['harga'], 'jumlah': qty, 'gambar': gambar}
                else:
                    # Cap at available stock
                    cart[nama] = {'harga': barang['harga'], 'jumlah': stok_tersedia, 'gambar': gambar}
        
        return redirect('/pembeli/keranjang')

    total = 0
    for nama in cart:
        total += cart[nama]['harga'] * cart[nama]['jumlah']
    return render_template('18-masukkankeranjang.html', cart=cart, total=total)

# ============================================================
# PEMBELI: PILIH PEMBAYARAN
# ============================================================

@app.route('/pembeli/pilih-pembayaran')
def pembeli_pilih_pembayaran():
    items = []
    total_int = 0
    for nama in cart:
        item = cart[nama]
        subtotal = item['harga'] * item['jumlah']
        total_int += subtotal
        items.append({'nama': nama, 'harga': item['harga'], 'qty': item['jumlah'], 'subtotal': subtotal, 'gambar': item.get('gambar', '')})
    total = total_int
    return render_template('34-pilihpembayaran.html', items=items, total=total, formatRp=formatRp)

# ============================================================
# PEMBELI: BAYAR TUNAI
# ============================================================

def get_items_bayar():
    items = []
    for nama in cart:
        items.append({"nama": nama, "jumlah": cart[nama]['jumlah'], "harga": cart[nama]['harga'], "diskon": 0})
    return items if items else [{"nama": "-", "jumlah": 0, "harga": 0, "diskon": 0}]

def buat_pesanan_dari_cart(metode):
    global cart
    if not cart:
        return
    trx_id = "TRX" + str(random.randint(10000, 99999))
    sekarang = datetime.now()
    tanggal = sekarang.strftime("%Y-%m-%d")
    waktu = sekarang.strftime("%H:%M:%S")
    pelanggan = session.get('nama') or session.get('user') or 'Guest'
    barang_list = []
    for nama in cart:
        qty = cart[nama]['jumlah']
        barang_list.append({"nama": nama, "jumlah": qty, "harga": cart[nama]['harga']})
        # Decrement stock in data_barang
        for b in data_barang:
            if b['nama'] == nama:
                b['stok'] = max(0, b['stok'] - qty)
                break
    save_data_barang(data_barang)
    
    pesanan_baru = {
        "id": trx_id,
        "tanggal": tanggal,
        "waktu": waktu,
        "pelanggan": pelanggan,
        "metode": metode,
        "status": "Disiapkan",
        "barang": barang_list,
        "total_awal": hitung_total_barang(barang_list),
        "total": hitung_total_barang(barang_list),
        "refund": 0
    }
    pesanan.append(pesanan_baru)
    save_pesanan(pesanan)
    cart = {}

@app.route('/pembeli/tunai')
def pembeli_tunai():
    session['metode'] = 'Tunai'
    items = get_items_bayar()
    buat_pesanan_dari_cart('Tunai')
    total = 0
    for item in items:
        total += item["jumlah"] * item["harga"]
    nama = session.get('nama') or session.get('user') or 'Guest'
    kode = random.randint(1000, 9999)
    session['tunai_items'] = items
    session['tunai_total'] = total
    session['tunai_kode'] = kode
    return render_template('11-rincian-tunai.html', items=items, total=total, nama=nama, kode=kode, status='belum')

# ============================================================
# PEMBELI: BAYAR QRIS
# ============================================================

@app.route('/pembeli/qris')
def pembeli_qris():
    session['metode'] = 'QRIS'
    if cart:
        subtotal = 0
        total_diskon = 0
        for item in get_items_bayar():
            subtotal += item["jumlah"] * item["harga"]
            total_diskon += item["diskon"]
        total = subtotal - total_diskon
        session['qris_total'] = total
        buat_pesanan_dari_cart('QRIS')
    else:
        total = session.get('qris_total')
        if total is None:
            pelanggan = session.get('nama') or session.get('user') or 'Guest'
            pesanan_user = [p for p in pesanan if p["pelanggan"] == pelanggan and p["metode"] == "QRIS"]
            if pesanan_user:
                total = pesanan_user[-1]["total"]
            else:
                total = 0
    return render_template('1-pembayaranqris.html', total=total, formatRp=formatRp)

# ============================================================
# PEMBELI: PESANAN

@app.route('/pembeli/selesai')
def pembeli_selesai():
    return redirect(url_for('pembeli_pesanan_selesai'))

# ============================================================

@app.route('/pembeli/pesanan')
def pembeli_pesanan():
    pelanggan = session.get('nama') or session.get('user') or 'Guest'
    pesanan_user = [p for p in pesanan if p["pelanggan"] == pelanggan]
    
    count_dikemas = len([p for p in pesanan_user if p["status"] == "Disiapkan"])
    count_siap = len([p for p in pesanan_user if p["status"] in ("Siap diambil", "Menunggu Pembayaran")])
    count_selesai = len([p for p in pesanan_user if p["status"] in ("Selesai", "Sudah diambil")])
    
    return render_template('8-lihatpesanan.html', 
                           count_dikemas=count_dikemas, 
                           count_siap=count_siap, 
                           count_selesai=count_selesai)

@app.route('/pembeli/pesanan/dikemas')
def pembeli_pesanan_dikemas():
    pelanggan = session.get('nama') or session.get('user') or 'Guest'
    pesanan_user = [p for p in pesanan if p["pelanggan"] == pelanggan]
    pesanan_to_show = [p for p in pesanan_user if p["status"] == "Disiapkan"]
    return render_template('8_1-detailpesanan.html', pesanan_list=pesanan_to_show)

@app.route('/pembeli/pesanan/siapdiambil')
def pembeli_pesanan_siapdiambil():
    pelanggan = session.get('nama') or session.get('user') or 'Guest'
    pesanan_user = [p for p in pesanan if p["pelanggan"] == pelanggan]
    pesanan_to_show = [p for p in pesanan_user if p["status"] in ("Siap diambil", "Menunggu Pembayaran")]
    return render_template('8_2-detailpesanan.html', pesanan_list=pesanan_to_show)

@app.route('/pembeli/pesanan/selesai')
def pembeli_pesanan_selesai():
    pelanggan = session.get('nama') or session.get('user') or 'Guest'
    pesanan_user = [p for p in pesanan if p["pelanggan"] == pelanggan]
    pesanan_to_show = [p for p in pesanan_user if p["status"] in ("Selesai", "Sudah diambil")]
    return render_template('8_3-detailpesanan.html', pesanan_list=pesanan_to_show)

@app.route('/pembeli/status')
def pembeli_status():
    return redirect(url_for('pembeli_pesanan_dikemas'))

@app.route("/siap-diambil")
def siap_diambil():
    return redirect(url_for('pembeli_pesanan_siapdiambil'))

# ============================================================
# PEMBELI: PENILAIAN
# ============================================================

produk_belum_dinilai = [
    {"id": 1, "nama": "Roti Aoka", "gambar": "gambar-dan-icon/gambar-roti-aoka.jpeg"},
    {"id": 2, "nama": "Air Mineral Ades", "gambar": "gambar-dan-icon/ades.jpg"},
    {"id": 3, "nama": "Bolpoin", "gambar": "gambar-dan-icon/bulpoin.jpg"},
]

penilaian_saya = [
    {"id": 1, "nama": "Roti Aoka", "gambar": "gambar-dan-icon/gambar-roti-aoka.jpeg", "rating": 4, "tanggal": "01-01-2026 11:24"},
    {"id": 2, "nama": "Air Mineral Ades", "gambar": "gambar-dan-icon/ades.jpg", "rating": 5, "tanggal": "20-12-2025 18:22"},
    {"id": 3, "nama": "Bolpoin", "gambar": "gambar-dan-icon/bulpoin.jpg", "rating": 3, "tanggal": "18-12-2025 11:35"},
]

@app.route("/pembeli/penilaian")
def pembeli_penilaian():
    return render_template("2-penilaianbarang.html", produk=produk_belum_dinilai, penilaian=penilaian_saya)

@app.route("/pembeli/rating")
def pembeli_rating():
    produk_id = request.args.get('id', type=int)
    nama_produk = "Roti Aoka"
    gambar_produk = "gambar-dan-icon/gambar-roti-aoka.jpeg"
    varian_produk = "Vanilla"
    tanggal_pembelian = "28 Nov 2025"
    nama_pengguna = session.get('nama', '')

    if produk_id:
        for p in produk_belum_dinilai:
            if p['id'] == produk_id:
                nama_produk = p['nama']
                gambar_produk = p['gambar']
                for b in data_barang:
                    if b['nama'].lower() in nama_produk.lower() or nama_produk.lower() in b['nama'].lower():
                        varian_produk = f"{b.get('berat', '')} {b.get('satuan', '')}".strip() or "Vanilla"
                        break
                break

    return render_template("4-inputpenilaian.html",
        nama_produk=nama_produk,
        gambar_produk=gambar_produk,
        varian_produk=varian_produk,
        tanggal_pembelian=tanggal_pembelian,
        nama_pengguna=nama_pengguna)

@app.route("/pembeli/submit-rating", methods=["POST"])
def pembeli_submit_rating():
    nama_produk = request.form.get('produk_nama', '')
    rating = int(request.form.get('rating', 0))
    ulasan = request.form.get('ulasan', '').strip()
    ulasan_tags = request.form.get('ulasan_tags', '').strip()
    nama = request.form.get('nama', '').strip()

    if not nama:
        nama = session.get('nama', 'Anonim')

    if rating < 1:
        rating = 1

    foto_filename = None
    foto = request.files.get('foto')
    if foto and foto.filename:
        foto_filename = secure_filename(foto.filename)
        upload_dir = os.path.join(BASE_DIR, 'static', 'uploads', 'rating')
        os.makedirs(upload_dir, exist_ok=True)
        foto.save(os.path.join(upload_dir, foto_filename))

    now = datetime.now()
    tanggal = now.strftime("%d-%m-%Y %H:%M")

    penilaian_saya.insert(0, {
        "id": len(penilaian_saya) + 1,
        "nama": nama_produk,
        "gambar": foto_filename or "default.jpg",
        "rating": rating,
        "tanggal": tanggal,
        "ulasan": ulasan,
        "ulasan_tags": ulasan_tags,
        "oleh": nama,
    })

    # Update rating on data_barang if matching
    for b in data_barang:
        if b['nama'] == nama_produk:
            b['rating'] = rating
            break

    return redirect(url_for('pembeli_penilaian'))

@app.route("/pembeli/like", methods=["POST"])
def pembeli_like():
    return redirect("/pembeli/penilaian")

# ============================================================
# PEMBELI: STRUK
# ============================================================

@app.route('/pembeli/struk')
def pembeli_struk():
    trx_id = request.args.get('trx_id')
    pelanggan = session.get('nama') or session.get('user') or 'Guest'
    
    if trx_id:
        order = next((p for p in pesanan if p["id"] == trx_id and p["pelanggan"] == pelanggan), None)
    else:
        pesanan_user = [p for p in pesanan if p["pelanggan"] == pelanggan]
        order = pesanan_user[-1] if pesanan_user else None
        
    if order:
        sumber = order["barang"]
        metode = order["metode"]
        total_final = order["total"]
        # Use order's date if available
        try:
            sekarang = datetime.strptime(order["tanggal"], "%Y-%m-%d")
        except:
            sekarang = datetime.now()
        trx_id = order["id"]
        
        if "waktu" in order:
            jam = order["waktu"]
        else:
            today_str = datetime.now().strftime("%Y-%m-%d")
            if order["tanggal"] == today_str:
                jam = datetime.now().strftime("%H:%M:%S")
            else:
                jam = "00:00:00"
    else:
        sumber = get_items_bayar()
        metode = session.get('metode', 'Tunai')
        total_final = None # Will calculate
        sekarang = datetime.now()
        trx_id = "TRX-TEMP"
        jam = sekarang.strftime("%H:%M:%S")

    daftar_produk = []
    for item in sumber:
        daftar_produk.append({
            "nama": item["nama"],
            "qty": item["jumlah"],
            "harga": item["harga"],
            "diskon": item.get("diskon", 0),
        })

    subtotal = 0
    total_diskon = 0
    total_produk = 0
    for item in daftar_produk:
        subtotal += item["qty"] * item["harga"]
        total_diskon += item["diskon"]
        total_produk += item["qty"]
    
    total = total_final if total_final is not None else (subtotal - total_diskon)

    if metode == "Tunai":
        tunai = 50000
        kembali = tunai - total
    else:
        tunai = total
        kembali = 0

    tanggal = format_kbbi_date(sekarang)
    kode = session.get('tunai_kode', random.randint(1000, 9999))

    return render_template('5-strukpembayaran.html', produk=daftar_produk, subtotal=subtotal, total_diskon=total_diskon, total=total, tunai=tunai, kembali=kembali, tanggal=tanggal, jam=jam, kode=kode, total_produk=total_produk, metode=metode, nama=pelanggan, trx_id=trx_id)


@app.route('/admin/cetak_pdf')
def admin_cetak_pdf():
    if not session.get('user'):
        return redirect(url_for('admin_login'))

    tanggal_awal = request.args.get('tanggal_awal', '')
    tanggal_akhir = request.args.get('tanggal_akhir', '')

    data_transaksi = []
    for p in pesanan:
        if (not tanggal_awal or p['tanggal'] >= tanggal_awal) and (not tanggal_akhir or p['tanggal'] <= tanggal_akhir):
            total_barang = sum(b['harga'] * b['jumlah'] for b in p['barang'])
            data_transaksi.append({
                'tanggal': p['tanggal'],
                'id': p['id'],
                'jumlah': sum(b['jumlah'] for b in p['barang']),
                'total': total_barang
            })

    if not data_transaksi and not (tanggal_awal or tanggal_akhir):
        data_transaksi = [
            {'tanggal': '2026-05-01', 'id': 'TRX001', 'jumlah': 3, 'total': 15000},
            {'tanggal': '2026-05-02', 'id': 'TRX002', 'jumlah': 5, 'total': 25000},
            {'tanggal': '2026-05-03', 'id': 'TRX003', 'jumlah': 2, 'total': 10000},
        ]

    total_pendapatan = sum(t['total'] for t in data_transaksi)
    modal_barang = int(total_pendapatan * 0.7)
    untung_rugi = total_pendapatan - modal_barang

    return render_template('12.-cetaklaporan_pdf.html',
        total_pendapatan=total_pendapatan,
        modal_barang=modal_barang,
        untung_rugi=untung_rugi,
        data_transaksi=data_transaksi)

@app.route('/admin/cetak_excel')
def admin_cetak_excel():
    if not session.get('user'):
        return redirect(url_for('admin_login'))

    tanggal_awal = request.args.get('tanggal_awal', '')
    tanggal_akhir = request.args.get('tanggal_akhir', '')

    data_transaksi = []
    for p in pesanan:
        if (not tanggal_awal or p['tanggal'] >= tanggal_awal) and (not tanggal_akhir or p['tanggal'] <= tanggal_akhir):
            total_barang = sum(b['harga'] * b['jumlah'] for b in p['barang'])
            data_transaksi.append({
                'tanggal': p['tanggal'],
                'id': p['id'],
                'jumlah': sum(b['jumlah'] for b in p['barang']),
                'total': total_barang
            })

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Penjualan"
    ws.append(['No', 'Tanggal', 'ID Transaksi', 'Jumlah Item', 'Total Pendapatan'])
    for i, t in enumerate(data_transaksi, 1):
        ws.append([i, format_kbbi_date(t['tanggal']), t['id'], t['jumlah'], formatRp(t['total'])])
    
    total_pendapatan = sum(t['total'] for t in data_transaksi)
    modal_barang = int(total_pendapatan * 0.7)
    untung_rugi = total_pendapatan - modal_barang

    ws.append([])
    ws.append(['Total Pendapatan', formatRp(total_pendapatan)])
    ws.append(['Modal Barang (70%)', formatRp(modal_barang)])
    ws.append(['Untung/Rugi', formatRp(untung_rugi)])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment;filename=laporan_penjualan.xlsx'}
    )

# ============================================================
# MAIN
# ============================================================

if __name__ == '__main__':
    import socket
    port = 5000
    while port <= 5010:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            if s.connect_ex(('127.0.0.1', port)) != 0:
                break
        port += 1
    print(f"Starting server on port {port}")
    app.run(debug=True, port=port)
