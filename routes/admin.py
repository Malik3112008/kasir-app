from flask import Blueprint, render_template, request, redirect, url_for, session, make_response, Response, jsonify
from datetime import datetime
import os
import secrets
import json
import random
import io
from werkzeug.utils import secure_filename
from database import db
from helpers import format_kbbi_date, formatRp

admin_bp = Blueprint('admin', __name__)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'static', 'gambar')

@admin_bp.route('/informasi')
def informasi():
    return render_template('informasi.html', koperasi=db.data_koperasi)

@admin_bp.route('/admin/denah')
def admin_denah():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    if session.get('role') != 'admin':
        return redirect(url_for('pembeli.pembeli_denah'))
    edit_mode = request.args.get('edit') == '1'
    edit_card_id = request.args.get('edit_card', type=int) if edit_mode else None
    edit_card = None
    if edit_card_id is not None:
        for card in db.CARDS_DATA:
            if card['id'] == edit_card_id:
                edit_card = card
                break
    return render_template('04.Denah.html', cards=db.CARDS_DATA, edit_mode=edit_mode, edit_card=edit_card, is_admin=True, page='home')

@admin_bp.route('/admin/delete/<int:card_id>', methods=['POST'])
def admin_delete_card(card_id):
    if not session.get('user'):
        return "Unauthorized", 403
    db.CARDS_DATA = [c for c in db.CARDS_DATA if c['id'] != card_id]
    return redirect(url_for('admin.admin_denah', edit=1))

@admin_bp.route('/admin/update/<int:card_id>', methods=['POST'])
def admin_update_card(card_id):
    if not session.get('user'):
        return "Unauthorized", 403
    for card in db.CARDS_DATA:
        if card['id'] == card_id:
            card['text'] = request.form.get('text', card['text'])
            card['icon'] = request.form.get('icon', card['icon'])
            card['icon_size'] = int(request.form.get('icon_size', card.get('icon_size', 30)))
            card['width'] = int(request.form.get('width', card['width']))
            card['height'] = int(request.form.get('height', card['height']))
            break
    return redirect(url_for('admin.admin_denah', edit=1))

@admin_bp.route('/admin/move/<int:card_id>', methods=['POST'])
def admin_move_card(card_id):
    if not session.get('user'):
        return "Unauthorized", 403
    data = request.get_json()
    for card in db.CARDS_DATA:
        if card['id'] == card_id:
            card['left'] = int(data.get('left', card['left']))
            card['top'] = int(data.get('top', card['top']))
            break
    return jsonify({'ok': True})

@admin_bp.route('/dynamic_cards.css')
def dynamic_cards_css():
    css_content = ""
    for card in db.CARDS_DATA:
        css_content += f".card-id-{card['id']} {{ width: {card['width']}px; height: {card['height']}px; left: {card['left']}px; top: {card['top']}px; }}\n"
    response = make_response(css_content)
    response.headers['Content-Type'] = 'text/css'
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return response

@admin_bp.route('/admin/denah/<folder>')
def admin_detail(folder):
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    if session.get('role') != 'admin':
        return redirect(url_for('pembeli.pembeli_detail', folder=folder))
    title = folder.replace('-', ' ').replace('_', ' ').title()
    return render_template('04.Denah.html', title=title, folder=folder, page='detail', is_admin=True)

@admin_bp.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        if db.USERS.get(username) == password:
            session['user'] = username
            session['role'] = 'admin'
            db.tambah_aktivitas("login", "Log in system", "Sukses", username)
            return redirect(url_for('admin.admin_dashboard'))
        error = 'Nama akun atau kata sandi tidak cocok.'
    return render_template('05.1.login.html', error=error)

@admin_bp.route('/admin/logout')
def admin_logout():
    admin_name = session.get('user', 'Admin')
    db.tambah_aktivitas("logout", "Log out system", "Sukses", admin_name)
    session.clear()
    return redirect(url_for('admin.admin_login'))


@admin_bp.route('/admin/register', methods=['GET', 'POST'])
def admin_register():
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        confirm = request.form.get('confirm', '')
        if not username or not email or not password:
            error = 'Lengkapi semua bidang.'
        elif password != confirm:
            error = 'Kata sandi dan konfirmasi tidak cocok.'
        elif username in db.USERS:
            error = 'Nama pengguna sudah terdaftar.'
        else:
            db.USERS[username] = password
            db.EMAIL_TO_USER[email] = username
            db.save_users()
            return redirect(url_for('admin.admin_login'))
    return render_template('05.2.register.html', error=error)


@admin_bp.route('/admin/forgot', methods=['GET', 'POST'])
def admin_forgot():
    error = None
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        session['reset_email'] = email
        session['reset_user'] = db.EMAIL_TO_USER.get(email)
        otp = f"{secrets.randbelow(900000) + 100000}"
        session['otp'] = otp
        print(f"[DEBUG] OTP for {email}: {otp}")
        return redirect(url_for('admin.admin_verify'))
    return render_template('05.3.forgot.html', error=error)

@admin_bp.route('/admin/verify', methods=['GET', 'POST'])
def admin_verify():
    error = None
    if request.method == 'POST':
        otp = request.form.get('otp', '').strip()
        if otp and otp == session.get('otp'):
            return redirect(url_for('admin.admin_reset'))
        error = 'Kode OTP tidak valid.'
    return render_template('05.4.verify_otp.html', error=error, email=session.get('reset_email'))

@admin_bp.route('/admin/reset', methods=['GET', 'POST'])
def admin_reset():
    error = None
    if request.method == 'POST':
        password = request.form.get('password', '')
        confirm = request.form.get('confirm', '')
        if not password:
            error = 'Masukkan kata sandi baru.'
        elif password != confirm:
            error = 'Konfirmasi kata sandi tidak cocok.'
        else:
            user = session.get('reset_user')
            if user:
                db.USERS[user] = password
                db.save_users()
            session.pop('otp', None)
            session.pop('reset_email', None)
            session.pop('reset_user', None)
            return redirect(url_for('admin.admin_login'))
    return render_template('05.5.reset.html', error=error)


@admin_bp.route("/admin/notifikasi")
def admin_notifikasi():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    return render_template("06.NotifikasiAdmin.html", notifikasi=db.notifikasi)

@admin_bp.route("/admin/riwayat")
def admin_riwayat():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    return render_template("06.RiwayatTransaksi.html", pesanan=db.pesanan)

@admin_bp.route("/admin/detail-transaksi/<trx_id>")
def admin_detail_transaksi(trx_id):
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    trx = None
    for p in db.pesanan:
        if p['id'] == trx_id:
            trx = p
            break
    if not trx:
        return "Transaksi tidak ditemukan", 404
    items = []
    for b in trx['barang']:
        items.append({
            'nama': b['nama'],
            'harga': b['harga'],
            'jumlah': b['jumlah'],
            'id_transaksi': trx['id'],
            'gambar': b.get('gambar', '')
        })
    total = sum(b['harga'] * b['jumlah'] for b in trx['barang'])
    tanggal_fmt = format_kbbi_date(trx['tanggal'])
    return render_template("detail_transaksi.html", items=items, total=total, tanggal=tanggal_fmt)

@admin_bp.route("/admin/riwayat-notifikasi")
def admin_riwayat_notifikasi():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    return render_template("06.RiwayatNotifikasi.html", riwayat=db.riwayat)

@admin_bp.route("/admin/riwayat-aktivitas")
def admin_riwayat_aktivitas():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    daftar_admin = list(set(d['admin'] for d in db.data_aktivitas))
    return render_template("24.RiwayatAktivitas.html", daftar_admin=daftar_admin)

@admin_bp.route("/admin/api/aktivitas")
def admin_api_aktivitas():
    search = request.args.get('search', '').lower()
    admin_filter = request.args.get('admin', 'Pilihan Admin')
    tanggal = request.args.get('tanggal', '')
    hasil = db.data_aktivitas
    if search:
        hasil = [d for d in hasil if search in d['catatan'].lower()]
    if admin_filter != 'Pilihan Admin':
        hasil = [d for d in hasil if admin_filter in d['admin']]
    if tanggal:
        hasil = [d for d in hasil if d['waktu'].startswith(tanggal)]
    return jsonify(hasil)

@admin_bp.route("/admin/hapus-semua", methods=["POST"])
def admin_hapus_semua():
    db.riwayat.extend(db.notifikasi)
    db.notifikasi.clear()
    db.save_notifikasi()
    return redirect(url_for('admin.admin_notifikasi'))

@admin_bp.route('/admin/hapus-notif/<int:index>', methods=['POST'])
def admin_hapus_notif(index):
    if index < len(db.notifikasi):
        db.riwayat.append(db.notifikasi[index])
        db.notifikasi.pop(index)
        db.save_notifikasi()
    return redirect(url_for('admin.admin_notifikasi'))

@admin_bp.route('/admin/hapus-riwayat-satuan/<int:index>', methods=['POST'])
def admin_hapus_riwayat_satuan(index):
    if index < len(db.riwayat):
        db.riwayat.pop(index)
        db.save_notifikasi()
    return redirect(url_for('admin.admin_riwayat'))

@admin_bp.route("/admin/hapus-riwayat", methods=["POST"])
def admin_hapus_riwayat():
    db.riwayat.clear()
    db.save_notifikasi()
    return redirect(url_for('admin.admin_riwayat'))


@admin_bp.route('/admin/dashboard')
def admin_dashboard():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))

    totalProduk = len(db.data_barang)
    stokMenipis = len([b for b in db.data_barang if b['stok'] <= 10])
    
    today_str = datetime.now().strftime("%Y-%m-%d")
    today_trx = [p for p in db.pesanan if p['tanggal'] == today_str]
    transaksiHariIni = len(today_trx) if today_trx else len(db.pesanan)
    
    total_income = sum(p['total'] for p in db.pesanan)
    pendapatan = formatRp(total_income)

    # Initialize dynamic hasil_penjualan dictionary
    hasil_penjualan = {
        "Senin": {"makanan": 0, "minuman": 0, "alat tulis": 0},
        "Selasa": {"makanan": 0, "minuman": 0, "alat tulis": 0},
        "Rabu": {"makanan": 0, "minuman": 0, "alat tulis": 0},
        "Kamis": {"makanan": 0, "minuman": 0, "alat tulis": 0},
        "Jumat": {"makanan": 0, "minuman": 0, "alat tulis": 0},
    }
    
    # Map product names to categories for quick lookup
    nama_kategori = {}
    for b in db.data_barang:
        nama_kategori[b['nama'].lower()] = b['kategori'].lower()
        
    hari_map = {
        0: "Senin",
        1: "Selasa",
        2: "Rabu",
        3: "Kamis",
        4: "Jumat"
    }
    
    for p in db.pesanan:
        try:
            dt = datetime.strptime(p['tanggal'], "%Y-%m-%d")
            wday = dt.weekday()
            if wday in hari_map:
                hari = hari_map[wday]
                for item in p.get('barang', []):
                    item_nama = item.get('nama', '').lower()
                    qty = item.get('jumlah', 0)
                    
                    kat = nama_kategori.get(item_nama, "")
                    if not kat:
                        if "roti" in item_nama or "mie" in item_nama or "donat" in item_nama or "keripik" in item_nama or "cemilan" in item_nama or "makanan" in item_nama:
                            kat = "makanan"
                        elif "air" in item_nama or "teh" in item_nama or "milk" in item_nama or "botol" in item_nama or "minuman" in item_nama:
                            kat = "minuman"
                        else:
                            kat = "alat tulis"
                            
                    if "makan" in kat:
                        kat_key = "makanan"
                    elif "minum" in kat:
                        kat_key = "minuman"
                    else:
                        kat_key = "alat tulis"
                        
                    hasil_penjualan[hari][kat_key] += qty
        except:
            continue

    # If no sales recorded at all, use default realistic data for visual demonstration
    total_sales_count = sum(sum(day.values()) for day in hasil_penjualan.values())
    if total_sales_count == 0:
        hasil_penjualan = {
            "Senin": {"makanan": 5, "minuman": 2, "alat tulis": 2},
            "Selasa": {"makanan": 0, "minuman": 5, "alat tulis": 5},
            "Rabu": {"makanan": 5, "minuman": 5, "alat tulis": 5},
            "Kamis": {"makanan": 6, "minuman": 7, "alat tulis": 5},
            "Jumat": {"makanan": 10, "minuman": 5, "alat tulis": 7},
        }

    y_hasilPenjualan = []
    for hari in hasil_penjualan:
        hasil = sum(hasil_penjualan[hari].values())
        y_hasilPenjualan.append(hasil)

    return render_template('07.Beranda_admin.html',
                           data_penjualan=y_hasilPenjualan,
                           card_product=totalProduk,
                           card_stock=stokMenipis,
                           card_transaction=transaksiHariIni,
                           card_income=pendapatan,
                           grafik_penjualan=hasil_penjualan)


@admin_bp.route('/admin/kelola_akun_penjual')
def admin_kelola_akun_penjual():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    
    keyword = request.args.get('cari', '').lower()
    if keyword:
        filtered = [p for p in db.data_penjual if keyword in p['nama'].lower() or keyword in p['email'].lower()]
    else:
        filtered = db.data_penjual
        
    return render_template('08.pengelola_akun_penjual.html', penjual=filtered, keyword=keyword)

@admin_bp.route('/admin/tambah-akun', methods=['GET', 'POST'])
def admin_tambah_akun():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    if request.method == 'POST':
        max_id = max([p['id'] for p in db.data_penjual], default=0)
        akun_baru = {
            'id': max_id + 1,
            'nama': request.form['nama'],
            'email': request.form['email'],
            'status': request.form['status'],
            'foto': 'profile.png'
        }
        db.data_penjual.append(akun_baru)
        db.save_penjual()
        return redirect(url_for('admin.admin_kelola_akun_penjual'))
    return render_template('08.tambah_akun.html')

@admin_bp.route('/admin/edit-akun/<int:id>', methods=['GET', 'POST'])
def admin_edit_akun(id):
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    
    akun = next((p for p in db.data_penjual if p['id'] == id), None)
    if not akun:
        return redirect(url_for('admin.admin_kelola_akun_penjual'))
        
    if request.method == 'POST':
        akun['nama'] = request.form['nama']
        akun['email'] = request.form['email']
        akun['status'] = request.form['status']
        db.save_penjual()
        return redirect(url_for('admin.admin_kelola_akun_penjual'))
    return render_template('08.edit_akun.html', akun=akun)

@admin_bp.route('/admin/hapus-akun/<int:id>')
def admin_hapus_akun(id):
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    db.data_penjual = [p for p in db.data_penjual if p['id'] != id]
    db.save_penjual()
    return redirect(url_for('admin.admin_kelola_akun_penjual'))


@admin_bp.route('/admin/manajemen-barang')
def admin_manajemen_barang():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    keyword = request.args.get('cari', '').lower()
    if keyword:
        filtered = [b for b in db.data_barang if keyword in b['nama'].lower() or keyword in b['kategori'].lower()]
    else:
        filtered = db.data_barang
    return render_template('09.manajemen_barang.html', data=filtered, keyword=request.args.get('cari', ''))

@admin_bp.route('/admin/pengisian_barang')
def admin_pengisian_barang():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    return render_template('10.pengisian_barang_.html')

@admin_bp.route('/admin/pengisian_barang/<int:id>')
def admin_pengisian_barang_restok(id):
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    barang = None
    for b in db.data_barang:
        if b['no'] == id:
            barang = b
            break
    if not barang:
        return redirect(url_for('admin.admin_pengisian_barang'))
    return render_template('10.pengisian_barang_.html', barang=barang)

@admin_bp.route('/admin/tambah-data-barang')
def admin_tambah_data_barang():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    return render_template('25.tambah_data_barang.html')

@admin_bp.route('/admin/simpan-barang-baru', methods=['POST'])
def admin_simpan_barang_baru():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    kategori = request.form.get('kategori', '')
    nama_barang = request.form.get('nama_barang', '')
    harga_beli = int(request.form.get('harga_beli', 0))
    harga_jual = int(request.form.get('harga_jual', 0))
    jumlah = int(request.form.get('jumlah', 0))
    tanggal = request.form.get('tanggal', '')
    variasi = request.form.get('variasi', '')
    volume = request.form.get('volume', '')
    rasa = request.form.get('rasa', '')
    expired = request.form.get('expired', '')
    deskripsi = request.form.get('deskripsi', '')
    gambar_path = ''
    file = request.files.get('gambar')
    if file and file.filename:
        filename = secure_filename(file.filename)
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        file.save(os.path.join(UPLOAD_FOLDER, filename))
        gambar_path = 'gambar/' + filename

    no_baru = max([b['no'] for b in db.data_barang], default=0) + 1
    db.data_barang.append({
        'no': no_baru, 'nama': nama_barang, 'berat': volume or '-',
        'stok': jumlah, 'harga': harga_jual, 'kategori': kategori,
        'tanggal': tanggal, 'gambar': gambar_path, 'rating': 0, 'emoji': '📦'
    })
    db.save_data_barang()
    
    admin_name = session.get('user', 'Admin')
    db.tambah_aktivitas("tambah", f"Menambahkan pilihan barang: {nama_barang}", "Berhasil", admin_name)
    
    return render_template('17.-konfirmasi-barang.html',
        nama_barang=nama_barang, kategori=kategori,
        harga_beli=harga_beli, harga_jual=harga_jual, 
        jumlah=jumlah, tanggal=tanggal, variasi=variasi,
        ukuran=volume, rasa=rasa, expired=expired,
        deskripsi=deskripsi, gambar=gambar_path)

@admin_bp.route('/admin/simpan', methods=['POST'])
def admin_simpan():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    kategori = request.form['kategori']
    nama_barang = request.form['nama_barang']
    tanggal = request.form.get('tanggal', '')
    jumlah = request.form['jumlah']
    harga_jual = request.form.get('harga_jual', request.form.get('harga', '0'))
    catatan = request.form.get('catatan', '')
    restok_id = request.form.get('restok_id')

    if restok_id:
        for b in db.data_barang:
            if str(b['no']) == str(restok_id):
                b['stok'] = b.get('stok', 0) + int(jumlah)
                db.save_data_barang()
                admin_name = session.get('user', 'Admin')
                db.tambah_aktivitas("restok", f"Melakukan restok produk: {b['nama']} (+{jumlah} unit)", "Berhasil", admin_name)
                break


    return render_template('10.rekap_barang.html', kategori=kategori, nama_barang=nama_barang, tanggal=tanggal, jumlah=jumlah, harga=harga_jual, catatan=catatan)

@admin_bp.route('/admin/konfirmasi-barang')
def admin_konfirmasi_barang():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    nama_barang = request.args.get('nama_barang', '-')
    kategori = request.args.get('kategori', '-')
    variasi = request.args.get('variasi', '-')
    ukuran = request.args.get('ukuran', '-')
    rasa = request.args.get('rasa', '-')
    expired = request.args.get('expired', '-')
    deskripsi = request.args.get('deskripsi', '-')
    harga_beli = request.args.get('harga_beli', '0')
    harga_jual = request.args.get('harga_jual', '0')
    jumlah = request.args.get('jumlah', '0')
    tanggal = request.args.get('tanggal', '-')
    
    return render_template('17.-konfirmasi-barang.html', 
        nama_barang=nama_barang, kategori=kategori, variasi=variasi,
        ukuran=ukuran, rasa=rasa, expired=expired, deskripsi=deskripsi,
        harga_beli=harga_beli, harga_jual=harga_jual, jumlah=jumlah, 
        tanggal=tanggal)

@admin_bp.route('/admin/stok-tersedia')
def admin_stok_tersedia():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    keyword = request.args.get('cari', '').lower()
    kategori = request.args.get('kategori', '')
    page = int(request.args.get('page', 1))
    per_page = 5

    filtered = db.data_barang
    if keyword:
        filtered = [b for b in filtered if keyword in b['nama'].lower()]
    if kategori:
        filtered = [b for b in filtered if b['kategori'] == kategori]

    total_halaman = max(1, (len(filtered) + per_page - 1) // per_page)
    start = (page - 1) * per_page
    data_page = filtered[start:start + per_page]

    return render_template('14.-stoktersedia.html',
        data=data_page, page=page, total_halaman=total_halaman,
        keyword=request.args.get('cari', ''), kategori=kategori)

@admin_bp.route('/admin/stok-tersedia/edit/<int:id>', methods=['GET', 'POST'])
def admin_stok_tersedia_edit(id):
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    barang = None
    for b in db.data_barang:
        if b['no'] == id:
            barang = b
            break
    if not barang:
        return "Barang tidak ditemukan", 404
    if request.method == 'POST':
        old_price = barang['harga']
        new_price = int(request.form.get('harga', barang['harga']))
        barang['nama'] = request.form.get('nama', barang['nama'])
        barang['berat'] = request.form.get('berat', barang['berat'])
        barang['kategori'] = request.form.get('kategori', barang['kategori'])
        barang['stok'] = int(request.form.get('stok', barang['stok']))
        barang['harga'] = new_price
        barang['satuan'] = request.form.get('satuan', barang.get('satuan', ''))
        barang['tanggal_restok'] = request.form.get('tanggal_restok', barang.get('tanggal_restok', ''))
        barang['expired'] = request.form.get('expired', barang.get('expired', ''))
        barang['alasan'] = request.form.get('alasan', barang.get('alasan', ''))
        db.save_data_barang()
        
        admin_name = session.get('user', 'Admin')
        if old_price != new_price:
            catatan = f"Mengubah harga: {barang['nama']} ({formatRp(old_price)} -> {formatRp(new_price)})"
        else:
            catatan = f"Mengubah data barang: {barang['nama']}"
        db.tambah_aktivitas("ubah", catatan, "Berhasil", admin_name)
        
        return redirect(url_for('admin.admin_stok_tersedia'))
    return render_template('14.-stoktersedia_edit.html', barang=barang)
 
@admin_bp.route('/admin/stok-tersedia/hapus/<int:id>', methods=['POST'])
def admin_stok_tersedia_hapus(id):
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    
    barang = None
    for b in db.data_barang:
        if b['no'] == id:
            barang = b
            break
            
    if barang:
        admin_name = session.get('user', 'Admin')
        db.tambah_aktivitas("hapus", f"Menghapus pilihan barang: {barang['nama']}", "Berhasil", admin_name)
        
    db.data_barang = [b for b in db.data_barang if b['no'] != id]
    db.save_data_barang()
    return redirect(url_for('admin.admin_stok_tersedia'))


@admin_bp.route('/admin/cetak_laporan', methods=['GET', 'POST'])
def admin_cetak_laporan():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))

    tanggal_awal = request.args.get('tanggal_awal', '')
    tanggal_akhir = request.args.get('tanggal_akhir', '')

    data_transaksi = []
    for p in db.pesanan:
        if (not tanggal_awal or p['tanggal'] >= tanggal_awal) and (not tanggal_akhir or p['tanggal'] <= tanggal_akhir):
            total_barang = sum(b['harga'] * b['jumlah'] for b in p['barang'])
            data_transaksi.append({
                'tanggal': p['tanggal'],
                'id': p['id'],
                'jumlah_item': sum(b['jumlah'] for b in p['barang']),
                'total': total_barang
            })

    if not data_transaksi and not (tanggal_awal or tanggal_akhir):
        data_transaksi = [
            {'tanggal': '2026-05-01', 'id': 'TRX001', 'jumlah_item': 3, 'total': 15000},
            {'tanggal': '2026-05-02', 'id': 'TRX002', 'jumlah_item': 5, 'total': 25000},
            {'tanggal': '2026-05-03', 'id': 'TRX003', 'jumlah_item': 2, 'total': 10000},
        ]

    total_pendapatan = sum(t['total'] for t in data_transaksi)
    modal_barang = int(total_pendapatan * 0.7)
    untung_rugi = total_pendapatan - modal_barang

    return render_template('12.-cetaklaporan.html',
        barang=db.data_barang, formatRp=formatRp,
        total_pendapatan=total_pendapatan,
        modal_barang=modal_barang,
        untung_rugi=untung_rugi,
        data_transaksi=data_transaksi,
        data_barang_list=db.data_barang,
        pesanan_list=db.pesanan,
        tanggal_awal=tanggal_awal,
        tanggal_akhir=tanggal_akhir)

@admin_bp.route('/admin/cetak_transaksi_pdf')
def admin_cetak_transaksi_pdf():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    tanggal_awal = request.args.get('tanggal_awal', '')
    tanggal_akhir = request.args.get('tanggal_akhir', '')
    filtered_pesanan = []
    for p in db.pesanan:
        if (not tanggal_awal or p['tanggal'] >= tanggal_awal) and (not tanggal_akhir or p['tanggal'] <= tanggal_akhir):
            filtered_pesanan.append(p)
    return render_template('12.-cetaklaporan_transaksi_pdf.html', pesanan=filtered_pesanan)

@admin_bp.route('/admin/cetak_transaksi_excel')
def admin_cetak_transaksi_excel():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    tanggal_awal = request.args.get('tanggal_awal', '')
    tanggal_akhir = request.args.get('tanggal_akhir', '')
    filtered_pesanan = []
    for p in db.pesanan:
        if (not tanggal_awal or p['tanggal'] >= tanggal_awal) and (not tanggal_akhir or p['tanggal'] <= tanggal_akhir):
            filtered_pesanan.append(p)
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Transaksi"
    ws.append(['No', 'Tanggal', 'ID Transaksi', 'Nama Produk', 'Total', 'Metode', 'Status'])
    for i, p in enumerate(filtered_pesanan, 1):
        total_barang = sum(b['harga'] * b['jumlah'] for b in p['barang'])
        nama_produk = ', '.join(b['nama'] for b in p['barang'])
        tgl_fmt = format_kbbi_date(p['tanggal'])
        ws.append([i, tgl_fmt, p['id'], nama_produk, formatRp(total_barang), p['metode'], p['status']])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment;filename=laporan_transaksi.xlsx'}
    )

@admin_bp.route('/admin/pengaturan-laporan')
def admin_pengaturan_laporan():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    tanggal_awal = request.args.get('tanggal_awal', '')
    tanggal_akhir = request.args.get('tanggal_akhir', '')
    laporan_type = request.args.get('laporan', 'barang')
    return render_template("16.pengaturan_laporan.html",
                           barang=db.data_barang,
                           tanggal_awal=tanggal_awal,
                           tanggal_akhir=tanggal_akhir,
                           laporan_type=laporan_type)

@admin_bp.route('/admin/cetak_laporan_barang')
def admin_cetak_laporan_barang():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    total_nilai = sum(b['harga'] * b['stok'] for b in db.data_barang)
    return render_template('12.-cetaklaporan_barang.html',
        data_barang=db.data_barang,
        total_nilai=total_nilai)

@admin_bp.route('/admin/cetak_barang_pdf')
def admin_cetak_barang_pdf():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    tanggal_awal = request.args.get('tanggal_awal', '')
    tanggal_akhir = request.args.get('tanggal_akhir', '')
    filtered_barang = []
    for b in db.data_barang:
        if (not tanggal_awal or b['tanggal'] >= tanggal_awal) and (not tanggal_akhir or b['tanggal'] <= tanggal_akhir):
            filtered_barang.append(b)
    total_nilai = sum(b['harga'] * b['stok'] for b in filtered_barang)
    return render_template('12.-cetaklaporan_barang_pdf.html',
        data_barang=filtered_barang,
        total_nilai=total_nilai)

@admin_bp.route('/admin/cetak_barang_excel')
def admin_cetak_barang_excel():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    tanggal_awal = request.args.get('tanggal_awal', '')
    tanggal_akhir = request.args.get('tanggal_akhir', '')
    filtered_barang = []
    for b in db.data_barang:
        if (not tanggal_awal or b['tanggal'] >= tanggal_awal) and (not tanggal_akhir or b['tanggal'] <= tanggal_akhir):
            filtered_barang.append(b)
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Data Barang"
    ws.append(['No', 'Nama', 'Stok', 'Harga', 'Kategori', 'Tanggal', 'Total'])
    for b in filtered_barang:
        ws.append([b['no'], b['nama'], b['stok'], formatRp(b['harga']), b['kategori'], format_kbbi_date(b['tanggal']), formatRp(b['harga'] * b['stok'])])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment;filename=laporan_data_barang.xlsx'}
    )

@admin_bp.route('/admin/cek-pembayaran')
def admin_cek_pembayaran():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    cari = request.args.get('cari', '').strip().lower()
    page = request.args.get('page', 1, type=int)
    per_page = 20
    if cari:
        hasil = [p for p in db.pesanan if cari in p['id'].lower() or cari in p['pelanggan'].lower()]
    else:
        hasil = list(db.pesanan)
    total = len(hasil)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    start = (page - 1) * per_page
    end = start + per_page
    transaksi = hasil[start:end]
    return render_template('21.cek_pembayaran.html', pesanan=transaksi,
                           page=page, total_pages=total_pages, total=total,
                           keyword=cari)

@admin_bp.route('/admin/cek-pembayaran/detail/<trx_id>')
def admin_cek_pembayaran_detail(trx_id):
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    order = None
    for p in db.pesanan:
        if p['id'] == trx_id:
            order = p
            break
    if not order:
        return "Transaksi tidak ditemukan", 404
    return render_template('21.cek_pembayaran_detail.html', order=order)

@admin_bp.route('/admin/cek-pembayaran/update-status/<trx_id>', methods=['POST'])
def admin_update_status(trx_id):
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    
    new_status = request.form.get('status', '').strip()
    if not new_status:
        return "Status tidak boleh kosong", 400
        
    order = None
    for p in db.pesanan:
        if p['id'] == trx_id:
            order = p
            break
            
    if not order:
        return "Transaksi tidak ditemukan", 404
        
    old_status = order.get('status', '')
    if old_status != new_status:
        order['status'] = new_status
        db.save_pesanan()
        
        if new_status == "Siap diambil":
            db.tambah_notifikasi(
                "Pembayaran Dikonfirmasi",
                f"Pembayaran {order['metode']} oleh {order['pelanggan']} senilai {formatRp(order['total'])} telah divalidasi",
                "hijau"
            )
        elif new_status == "Selesai":
            db.tambah_notifikasi(
                "Transaksi Selesai",
                f"Transaksi #{trx_id} oleh {order['pelanggan']} senilai {formatRp(order['total'])} telah selesai",
                "hijau"
            )
            
        admin_name = session.get('user', 'Admin')
        db.tambah_aktivitas(
            "ubah", 
            f"Mengubah status transaksi #{trx_id}: {old_status} -> {new_status}", 
            "Berhasil", 
            admin_name
        )
        
    return redirect(url_for('admin.admin_cek_pembayaran_detail', trx_id=trx_id))


@admin_bp.route('/admin/pengaturan', methods=['GET'])
def admin_pengaturan():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    
    logo_path = db.data_koperasi.get('logo', '')
    if logo_path:
        full_path = os.path.join(BASE_DIR, 'static', logo_path)
        if not os.path.exists(full_path):
            logo_path = 'image/logo_3.png'
    else:
        logo_path = 'image/logo_3.png'
        
    data = dict(db.data_koperasi)
    data['logo'] = logo_path
    return render_template('22.pengaturan_umum.html', **data)

@admin_bp.route('/admin/simpan_pengaturan', methods=['POST'])
def admin_simpan_pengaturan():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))
    db.data_koperasi['nama'] = request.form.get('nama', db.data_koperasi['nama'])
    db.data_koperasi['deskripsi'] = request.form.get('deskripsi', db.data_koperasi['deskripsi'])
    db.data_koperasi['alamat'] = request.form.get('alamat', db.data_koperasi['alamat'])
    db.data_koperasi['telepon'] = request.form.get('telepon', db.data_koperasi['telepon'])
    db.data_koperasi['jam'] = request.form.get('jam', db.data_koperasi['jam'])
    db.data_koperasi['hari'] = request.form.get('hari', db.data_koperasi['hari'])

    file = request.files.get('gambar')
    if file and file.filename:
        filename = secure_filename(file.filename)
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        file.save(os.path.join(UPLOAD_FOLDER, filename))
        db.data_koperasi['logo'] = 'gambar/' + filename
        
        file.seek(0)
        file.save(os.path.join(BASE_DIR, 'static', 'image', 'logo_3.png'))

    db.save_koperasi()
    return redirect(url_for('admin.admin_pengaturan'))

@admin_bp.route('/admin/cetak_pdf')
def admin_cetak_pdf():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))

    tanggal_awal = request.args.get('tanggal_awal', '')
    tanggal_akhir = request.args.get('tanggal_akhir', '')

    data_transaksi = []
    for p in db.pesanan:
        if (not tanggal_awal or p['tanggal'] >= tanggal_awal) and (not tanggal_akhir or p['tanggal'] <= tanggal_akhir):
            total_barang = sum(b['harga'] * b['jumlah'] for b in p['barang'])
            data_transaksi.append({
                'tanggal': p['tanggal'],
                'id': p['id'],
                'jumlah': sum(b['jumlah'] for b in p['barang']),
                'total': total_barang
            })

    if not data_transaksi and not (tanggal_awal or tanggal_akhir):
        data_transaksi = [
            {'tanggal': '2026-05-01', 'id': 'TRX001', 'jumlah': 3, 'total': 15000},
            {'tanggal': '2026-05-02', 'id': 'TRX002', 'jumlah': 5, 'total': 25000},
            {'tanggal': '2026-05-03', 'id': 'TRX003', 'jumlah': 2, 'total': 10000},
        ]

    total_pendapatan = sum(t['total'] for t in data_transaksi)
    modal_barang = int(total_pendapatan * 0.7)
    untung_rugi = total_pendapatan - modal_barang

    return render_template('12.-cetaklaporan_pdf.html',
        total_pendapatan=total_pendapatan,
        modal_barang=modal_barang,
        untung_rugi=untung_rugi,
        data_transaksi=data_transaksi)

@admin_bp.route('/admin/cetak_excel')
def admin_cetak_excel():
    if not session.get('user'):
        return redirect(url_for('admin.admin_login'))

    tanggal_awal = request.args.get('tanggal_awal', '')
    tanggal_akhir = request.args.get('tanggal_akhir', '')

    data_transaksi = []
    for p in db.pesanan:
        if (not tanggal_awal or p['tanggal'] >= tanggal_awal) and (not tanggal_akhir or p['tanggal'] <= tanggal_akhir):
            total_barang = sum(b['harga'] * b['jumlah'] for b in p['barang'])
            data_transaksi.append({
                'tanggal': p['tanggal'],
                'id': p['id'],
                'jumlah': sum(b['jumlah'] for b in p['barang']),
                'total': total_barang
            })

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Penjualan"
    ws.append(['No', 'Tanggal', 'ID Transaksi', 'Jumlah Item', 'Total Pendapatan'])
    for i, t in enumerate(data_transaksi, 1):
        ws.append([i, format_kbbi_date(t['tanggal']), t['id'], t['jumlah'], formatRp(t['total'])])
    
    total_pendapatan = sum(t['total'] for t in data_transaksi)
    modal_barang = int(total_pendapatan * 0.7)
    untung_rugi = total_pendapatan - modal_barang

    ws.append([])
    ws.append(['Total Pendapatan', formatRp(total_pendapatan)])
    ws.append(['Modal Barang (70%)', formatRp(modal_barang)])
    ws.append(['Untung/Rugi', formatRp(untung_rugi)])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment;filename=laporan_penjualan.xlsx'}
    )
